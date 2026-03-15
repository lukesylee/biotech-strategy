"""
Pipeline Enrichment Script
===========================
Enriches biotech_universe_raw.csv with pipeline data from ClinicalTrials.gov API.
Run AFTER build_universe.py has generated biotech_universe_raw.csv.

Requirements:
    pip install requests pandas openpyxl tqdm

Usage:
    python build_pipeline.py

Output:
    pipeline_db.csv                    (all trials per company)
    biotech_universe_enriched.csv      (universe with lead asset / stage added)
    Biotech_Stock_Screener_FULL.xlsx   (updated with pipeline data)
"""

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import json
from datetime import datetime

UNIVERSE_CSV  = "biotech_universe_raw.csv"
PIPELINE_CSV  = "pipeline_db.csv"
ENRICHED_CSV  = "biotech_universe_enriched.csv"
SCREENER_XLSX = "Biotech_Stock_Screener_FULL.xlsx"
HEADERS       = {"User-Agent": "biotech-screener research@example.com"}

# Phase → Dev Stage mapping
PHASE_MAP = {
    "PHASE1":  "Phase 1",
    "PHASE2":  "Phase 2",
    "PHASE3":  "Phase 3",
    "PHASE1|PHASE2": "Phase 1/2",
    "PHASE2|PHASE3": "Phase 2/3",
    "NA":      "Preclinical",
    "EARLY_PHASE1": "Phase 1",
}

# Therapy area keyword mapping (applied to condition/indication text)
THERAPY_AREA_MAP = [
    ("Oncology",        ["cancer", "tumor", "tumour", "carcinoma", "lymphoma", "leukemia",
                         "melanoma", "sarcoma", "glioma", "glioblastoma", "myeloma",
                         "oncology", "neoplasm", "malignant"]),
    ("Rare Disease",    ["rare", "orphan", "duchenne", "pompe", "fabry", "gaucher",
                         "wilson", "huntington", "amyloid", "ttr", "mps", "hemophilia",
                         "sickle cell", "thalassemia", "cystic fibrosis", "spinal muscular"]),
    ("CNS",             ["alzheimer", "parkinson", "epilepsy", "seizure", "multiple sclerosis",
                         "schizophrenia", "depression", "anxiety", "rett", "autism",
                         "amyotrophic", "als", "neuropathy", "migraine", "tremor",
                         "neurological", "psychiatric", "cns", "brain"]),
    ("Immunology",      ["autoimmune", "rheumatoid", "lupus", "psoriasis", "crohn",
                         "colitis", "ibd", "sjogren", "myasthenia", "cidp", "itp",
                         "pemphigus", "immunology", "inflammation", "atopic", "eczema",
                         "asthma", "ankylosing"]),
    ("Cardiometabolic", ["heart failure", "cardiac", "cardiovascular", "atherosclerosis",
                         "hypertension", "cholesterol", "ldl", "triglyceride", "diabetes",
                         "obesity", "nash", "mash", "nafld", "metabolic", "lipid",
                         "coronary", "myocardial"]),
    ("Infectious Disease",["hiv", "hepatitis", "tuberculosis", "malaria", "influenza",
                           "covid", "sars", "viral", "bacterial", "fungal", "antibiotic",
                           "antimicrobial", "infection"]),
]


def get_therapy_area(conditions: list) -> str:
    """Classify therapy area from a list of condition strings."""
    text = " ".join(str(c).lower() for c in conditions)
    for area, keywords in THERAPY_AREA_MAP:
        if any(kw in text for kw in keywords):
            return area
    return "Other"


def get_highest_phase(phase_list: list) -> str:
    """Return the highest development phase from a list."""
    priority = ["PHASE3", "PHASE2|PHASE3", "PHASE2", "PHASE1|PHASE2", "PHASE1", "EARLY_PHASE1", "NA"]
    for p in priority:
        if p in [str(x).upper() for x in phase_list]:
            return PHASE_MAP.get(p, "Preclinical")
    return "Preclinical"


def fetch_trials_for_company(company_name: str) -> list:
    """
    Fetch all clinical trials from ClinicalTrials.gov API v2 for a given sponsor.
    Returns list of trial dicts.
    """
    url = "https://clinicaltrials.gov/api/v2/studies"
    params = {
        "query.spons": company_name,
        "filter.overallStatus": "RECRUITING,ACTIVE_NOT_RECRUITING,COMPLETED,ENROLLING_BY_INVITATION",
        "fields": ("NCTId,BriefTitle,OfficialTitle,Phase,OverallStatus,"
                   "LeadSponsorName,Condition,InterventionType,InterventionName,"
                   "PrimaryCompletionDate,StudyFirstPostDate,StartDate"),
        "pageSize": 50,
        "format": "json",
    }
    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return []
        data = r.json()
        studies = data.get("studies", [])
        return studies
    except Exception as e:
        return []


def parse_trial(study: dict, ticker: str, company: str) -> dict:
    """Parse a ClinicalTrials.gov study dict into a flat row for the pipeline DB."""
    proto = study.get("protocolSection", {})
    id_mod     = proto.get("identificationModule", {})
    status_mod = proto.get("statusModule", {})
    design_mod = proto.get("designModule", {})
    conds_mod  = proto.get("conditionsModule", {})
    arms_mod   = proto.get("armsInterventionsModule", {})
    sponsor_mod= proto.get("sponsorCollaboratorsModule", {})

    nct_id     = id_mod.get("nctId", "")
    title      = id_mod.get("briefTitle", "") or id_mod.get("officialTitle", "")
    phase_list = design_mod.get("phases", ["NA"])
    conditions = conds_mod.get("conditions", [])
    interventions = arms_mod.get("interventions", [])
    
    # Get intervention name (drug name)
    drug_names = [i.get("name", "") for i in interventions if i.get("type") in 
                  ["DRUG", "BIOLOGICAL", "GENETIC", "COMBINATION_PRODUCT", "DEVICE"]]
    drug_name  = drug_names[0] if drug_names else ""
    
    # Modality from intervention type
    int_types  = [i.get("type", "") for i in interventions]
    if "GENETIC" in int_types:
        modality = "Gene Therapy / Gene Editing"
    elif "BIOLOGICAL" in int_types:
        modality = "Monoclonal Ab (mAb)"  # Rough — biologics are mainly mAbs
    elif "DRUG" in int_types:
        modality = "Small Molecule"
    else:
        modality = ""

    phase_str  = get_highest_phase(phase_list)
    therapy    = get_therapy_area(conditions)
    
    prim_compl = (status_mod.get("primaryCompletionDateStruct") or {}).get("date", "")
    overall_st = status_mod.get("overallStatus", "")
    
    # Collaborators = potential partners
    collabs    = sponsor_mod.get("collaborators", [])
    partners   = "; ".join(c.get("name","") for c in collabs if c.get("name","") != company)

    return {
        "Ticker":          ticker,
        "Company":         company,
        "NCT_ID":          nct_id,
        "Trial Title":     title[:120],
        "Drug / Asset":    drug_name[:60],
        "Modality":        modality,
        "Therapy Area":    therapy,
        "Indication":      "; ".join(conditions[:3]),
        "Phase":           phase_str,
        "Status":          overall_st,
        "Primary Compl.":  prim_compl,
        "Partners":        partners[:80],
    }


def enrich_pipeline(universe_csv: str) -> tuple:
    """
    For each company in the universe CSV, fetch trials from ClinicalTrials.gov.
    Returns (pipeline_df, enriched_universe_df).
    """
    print(f"\n[1/3] Loading universe from {universe_csv}...")
    univ = pd.read_csv(universe_csv)
    print(f"   {len(univ)} companies to process")

    all_trials = []
    company_summary = []

    for i, (_, row) in enumerate(univ.iterrows()):
        ticker  = row["Ticker"]
        company = row["Company Name"]

        if i % 20 == 0:
            print(f"   Progress: {i}/{len(univ)} — {ticker}")

        trials = fetch_trials_for_company(company)

        if trials:
            parsed = [parse_trial(t, ticker, company) for t in trials]
            all_trials.extend(parsed)

            # Determine lead asset = highest phase trial
            phase_priority = {"Phase 3": 6, "Phase 2/3": 5, "Phase 2": 4,
                               "Phase 1/2": 3, "Phase 1": 2, "Preclinical": 1, "": 0}
            parsed_sorted = sorted(parsed, key=lambda x: phase_priority.get(x["Phase"], 0), reverse=True)
            lead = parsed_sorted[0] if parsed_sorted else {}

            company_summary.append({
                "Ticker":          ticker,
                "Therapy Area":    lead.get("Therapy Area", ""),
                "Modality":        lead.get("Modality", ""),
                "Lead_Asset":      lead.get("Drug / Asset", ""),
                "Dev_Stage":       lead.get("Phase", ""),
                "Sub_Indication":  lead.get("Indication", "")[:60],
                "Partners":        lead.get("Partners", ""),
                "Trial_Count":     len(parsed),
            })
        else:
            company_summary.append({
                "Ticker": ticker, "Therapy Area": "", "Modality": "",
                "Lead_Asset": "", "Dev_Stage": "", "Sub_Indication": "",
                "Partners": "", "Trial_Count": 0,
            })

        time.sleep(0.2)  # Polite rate limit

    pipeline_df  = pd.DataFrame(all_trials)
    summary_df   = pd.DataFrame(company_summary)
    enriched_df  = univ.merge(summary_df, on="Ticker", how="left", suffixes=("", "_ct"))

    # Fill blanks in Therapy Area / Modality from ClinicalTrials data
    for col in ["Therapy Area", "Modality", "Lead_Asset", "Dev_Stage",
                "Sub_Indication", "Partners", "Trial_Count"]:
        if f"{col}_ct" in enriched_df.columns:
            enriched_df[col] = enriched_df[col].fillna("").astype(str)
            enriched_df[f"{col}_ct"] = enriched_df[f"{col}_ct"].fillna("").astype(str)
            mask = enriched_df[col] == ""
            enriched_df.loc[mask, col] = enriched_df.loc[mask, f"{col}_ct"]
            enriched_df.drop(columns=[f"{col}_ct"], inplace=True)

    print(f"\n   Total trials fetched: {len(pipeline_df)}")
    print(f"   Companies with trial data: {(summary_df['Trial_Count'] > 0).sum()}")
    return pipeline_df, enriched_df


def write_pipeline_sheet(wb: openpyxl.Workbook, pipeline_df: pd.DataFrame):
    """Write pipeline data to the 🧬 Pipeline DB sheet."""
    ws = wb["🧬 Pipeline DB"]
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),  right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),   bottom=Side(style="thin", color="CCCCCC"),
    )
    # Clear existing data rows
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for cell in row: cell.value = None

    FILL_W = PatternFill("solid", fgColor="FFFFFF")
    FILL_T = PatternFill("solid", fgColor="E6F4F6")

    # Map pipeline_df columns to sheet columns
    col_map = {
        1:  "Ticker",       2: "Company",       3: "Drug / Asset",
        4:  "Drug / Asset", 5: "Modality",       6: "Therapy Area",
        7:  "Indication",   8: "Phase",          9: "NCT_ID",
        10: "Primary Compl.", 11: "Partners",    12: "Status",
    }

    for row_idx, (_, row) in enumerate(pipeline_df.iterrows(), start=6):
        fill = FILL_W if row_idx % 2 == 0 else FILL_T
        ws.row_dimensions[row_idx].height = 28
        for col, key in col_map.items():
            c = ws.cell(row=row_idx, column=col)
            c.value = str(row.get(key, ""))[:100]
            c.font = Font(name="Arial", size=9, bold=(col == 1))
            c.fill = fill
            c.border = thin
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col in [2, 3, 7, 11]:
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    ws.auto_filter.ref = f"A5:{get_column_letter(18)}{5 + len(pipeline_df)}"


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import os
    print("=" * 60)
    print("  PIPELINE ENRICHMENT — ClinicalTrials.gov")
    print(f"  {datetime.today().strftime('%d %b %Y %H:%M')}")
    print("=" * 60)

    if not os.path.exists(UNIVERSE_CSV):
        print(f"❌ {UNIVERSE_CSV} not found. Run build_universe.py first.")
        exit(1)

    if not os.path.exists(SCREENER_XLSX):
        print(f"❌ {SCREENER_XLSX} not found. Run build_universe.py first.")
        exit(1)

    # Enrich with ClinicalTrials.gov
    pipeline_df, enriched_df = enrich_pipeline(UNIVERSE_CSV)

    # Save CSVs
    pipeline_df.to_csv(PIPELINE_CSV, index=False)
    enriched_df.to_csv(ENRICHED_CSV, index=False)
    print(f"\n[2/3] Saved: {PIPELINE_CSV} ({len(pipeline_df)} trials)")
    print(f"       Saved: {ENRICHED_CSV} ({len(enriched_df)} companies)")

    # Update Excel screener with pipeline data
    print(f"\n[3/3] Updating {SCREENER_XLSX} with pipeline data...")
    wb = openpyxl.load_workbook(SCREENER_XLSX)
    write_pipeline_sheet(wb, pipeline_df)
    wb.save(SCREENER_XLSX)

    print("\n✅ Pipeline enrichment complete!")
    print(f"   Updated screener: {SCREENER_XLSX}")
    print(f"\n   Therapy area coverage:")
    if "Therapy Area" in enriched_df.columns:
        print(enriched_df["Therapy Area"].value_counts().to_string())
    print(f"\n   Dev Stage breakdown:")
    if "Dev_Stage" in enriched_df.columns:
        print(enriched_df["Dev_Stage"].value_counts().to_string())
