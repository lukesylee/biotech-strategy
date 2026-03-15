"""
Biotech Universe Builder — SIC 2836
=====================================
Run this script LOCALLY (not in Claude's sandbox) to:
  1. Pull all SIC-2836 tickers from SEC EDGAR company facts
  2. Enrich with market data via yFinance
  3. Populate the Biotech_Stock_Screener.xlsx

Requirements:
    pip install requests pandas yfinance openpyxl tqdm

Usage:
    python build_universe.py

Output:
    Biotech_Stock_Screener_FULL.xlsx   (populated screener)
    biotech_universe_raw.csv           (raw ticker list for inspection)
"""

import requests
import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import json
import os
from datetime import datetime

# ── Configuration ─────────────────────────────────────────────────────────────
SIC_CODE        = "2836"
EXCHANGES       = {"NASDAQ", "NYSE", "NYSE MKT", "NYSE AMERICAN"}  # NYSE MKT = NYSE American
MIN_MARKET_CAP  = 50_000_000      # $50M floor
MAX_WORKERS     = 1               # yFinance rate limit: keep at 1 to avoid bans
BATCH_PAUSE_SEC = 0.5             # seconds between yFinance calls
INPUT_XLSX      = "Biotech_Stock_Screener.xlsx"   # your existing screener
OUTPUT_XLSX     = "Biotech_Stock_Screener_FULL.xlsx"
RAW_CSV         = "biotech_universe_raw.csv"

HEADERS = {"User-Agent": "biotech-screener research@example.com"}  # required by SEC


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — Pull SIC 2836 company list from SEC EDGAR
# ══════════════════════════════════════════════════════════════════════════════

def fetch_sec_sic_companies(sic_code: str) -> pd.DataFrame:
    """
    Fetches all companies with a given SIC code from SEC EDGAR.
    Returns DataFrame with columns: cik, name, ticker, exchange, sic.
    
    API docs: https://efts.sec.gov/LATEST/search-index
    Company tickers JSON: https://www.sec.gov/files/company_tickers_exchange.json
    """
    print(f"\n[1/4] Fetching SIC {sic_code} universe from SEC EDGAR...")

    # The most reliable source: SEC's full company→ticker→exchange mapping
    url = "https://www.sec.gov/files/company_tickers_exchange.json"
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    data = r.json()

    # data["data"] is a list of [cik, name, ticker, exchange]
    df = pd.DataFrame(data["data"], columns=["cik", "name", "ticker", "exchange"])
    print(f"   Total SEC-registered companies: {len(df):,}")

    # Now fetch the SIC code for each CIK using the company facts endpoint
    # To avoid hammering the API, use the bulk submissions dataset instead
    sic_url = "https://www.sec.gov/files/company_tickers.json"
    r2 = requests.get(sic_url, headers=HEADERS, timeout=30)
    r2.raise_for_status()
    tickers_data = r2.json()

    # Build CIK → SIC lookup via individual company submissions
    # For bulk SIC filtering, use the EDGAR full-text search company endpoint
    sic_search_url = (
        f"https://efts.sec.gov/LATEST/search-index?"
        f"q=%22SIC%3D{sic_code}%22&forms=10-K,10-Q&dateRange=custom"
        f"&startdt=2022-01-01&enddt=2024-12-31"
    )

    # Alternative: use the company search API which supports SIC filtering
    # This is the most direct and reliable approach
    sic_companies = []
    page = 0
    while True:
        api_url = (
            f"https://efts.sec.gov/LATEST/search-index?"
            f"q=%22%22&forms=10-K&dateRange=custom&startdt=2023-01-01"
            f"&enddt=2024-12-31&_source=file_date,period_of_report,entity_name,"
            f"file_num,period_of_report,biz_location,inc_states,category"
            f"&from={page * 100}&size=100"
        )
        # Use the EDGAR company search with SIC parameter
        comp_url = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&SIC={sic_code}&dateb=&owner=include&count=100&search_text=&action=getcompany&start={page*100}&output=atom"
        
        try:
            resp = requests.get(comp_url, headers=HEADERS, timeout=20)
            if resp.status_code != 200:
                break
            # Parse the atom feed for CIKs
            import xml.etree.ElementTree as ET
            ns = {"atom": "http://www.w3.org/2005/Atom"}
            root = ET.fromstring(resp.text)
            entries = root.findall("atom:entry", ns)
            if not entries:
                break
            for entry in entries:
                cik_el = entry.find("atom:cik-number", ns) or entry.find("cik-number")
                name_el = entry.find("atom:company-name", ns) or entry.find("company-name")
                if cik_el is not None and name_el is not None:
                    sic_companies.append({
                        "cik": cik_el.text.strip().lstrip("0"),
                        "sec_name": name_el.text.strip()
                    })
            print(f"   Page {page+1}: {len(entries)} companies fetched (total so far: {len(sic_companies)})")
            if len(entries) < 100:
                break
            page += 1
            time.sleep(0.3)
        except Exception as e:
            print(f"   Warning: page {page} failed — {e}")
            break

    if not sic_companies:
        print("   EDGAR atom feed unavailable. Falling back to company_tickers cross-reference...")
        # Fallback: pull individual CIK submissions to check SIC
        # This is slower but works if the atom feed is blocked
        return fetch_sic_via_submissions(df, sic_code)

    sic_df = pd.DataFrame(sic_companies).drop_duplicates("cik")
    sic_df["cik"] = sic_df["cik"].astype(str)
    df["cik"] = df["cik"].astype(str)

    # Merge to get tickers for SIC 2836 companies
    merged = sic_df.merge(df, on="cik", how="left")
    merged = merged.dropna(subset=["ticker"])
    merged = merged[merged["exchange"].str.upper().isin({e.upper() for e in EXCHANGES})]
    print(f"   SIC {sic_code} companies with listed tickers on target exchanges: {len(merged)}")
    return merged


def fetch_sic_via_submissions(df: pd.DataFrame, sic_code: str) -> pd.DataFrame:
    """
    Fallback: checks individual company submissions for SIC code.
    Uses only the tickers we already have from company_tickers_exchange.json
    and samples their CIK→SIC from EDGAR submissions.
    """
    print("   Using submissions fallback (this takes a few minutes)...")
    
    # Filter to exchanges we care about first
    candidates = df[df["exchange"].str.upper().isin({e.upper() for e in EXCHANGES})].copy()
    print(f"   Checking {len(candidates)} listed companies for SIC {sic_code}...")
    
    sic_matches = []
    for i, (_, row) in enumerate(candidates.iterrows()):
        cik_padded = str(row["cik"]).zfill(10)
        sub_url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
        try:
            r = requests.get(sub_url, headers=HEADERS, timeout=10)
            if r.status_code == 200:
                sub = r.json()
                if str(sub.get("sic", "")) == sic_code:
                    sic_matches.append(row)
        except:
            pass
        if i % 100 == 0 and i > 0:
            print(f"   Checked {i}/{len(candidates)}...")
        time.sleep(0.1)  # Be polite to SEC
    
    result = pd.DataFrame(sic_matches)
    print(f"   Found {len(result)} SIC {sic_code} companies via submissions")
    return result


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Enrich with market data via yFinance
# ══════════════════════════════════════════════════════════════════════════════

def enrich_with_yfinance(tickers: list) -> pd.DataFrame:
    """
    Pulls key financial metrics for each ticker from yFinance.
    Returns DataFrame with one row per company.
    """
    print(f"\n[2/4] Enriching {len(tickers)} tickers with yFinance data...")
    print("      (This will take several minutes — ~0.5s per ticker)")

    rows = []
    failed = []

    for i, ticker in enumerate(tickers):
        if i % 50 == 0:
            print(f"   Progress: {i}/{len(tickers)} ({100*i//len(tickers)}%)")
        try:
            t = yf.Ticker(ticker)
            info = t.info

            mkt_cap = info.get("marketCap", 0) or 0
            if mkt_cap < MIN_MARKET_CAP:
                continue  # Skip sub-$50M companies

            # Financials: cash + burn from balance sheet / cash flow
            cash = (info.get("totalCash") or 0) / 1e6  # Convert to $M
            
            # Quarterly burn: use operating cash flow (negative = burning)
            try:
                cf = t.quarterly_cashflow
                if cf is not None and not cf.empty and "Operating Cash Flow" in cf.index:
                    # Average of last 2 quarters
                    op_cf = cf.loc["Operating Cash Flow"].iloc[:2].mean()
                    burn = max(-op_cf / 1e6, 0)  # Positive = burning cash
                else:
                    burn = (info.get("operatingCashflow") or 0)
                    burn = max(-burn / 1e6 / 4, 0)  # Annual → quarterly
            except:
                burn = 0

            # 52-week high/low
            hi52 = info.get("fiftyTwoWeekHigh") or 0
            lo52 = info.get("fiftyTwoWeekLow") or 0

            rows.append({
                "Ticker":           ticker,
                "Company Name":     info.get("longName") or info.get("shortName") or ticker,
                "Exchange":         info.get("exchange") or "",
                "Market Cap ($M)":  round(mkt_cap / 1e6, 0),
                "Cash ($M)":        round(cash, 1),
                "Qtrly Burn ($M)":  round(burn, 1),
                "52W High ($)":     round(hi52, 2),
                "52W Low ($)":      round(lo52, 2),
                "Share Price ($)":  round(info.get("currentPrice") or info.get("previousClose") or 0, 2),
                "Shares Out (M)":   round((info.get("sharesOutstanding") or 0) / 1e6, 1),
                "Sector":           info.get("sector") or "",
                "Industry":         info.get("industry") or "",
                "Website":          info.get("website") or "",
                "Description":      (info.get("longBusinessSummary") or "")[:300],
            })
        except Exception as e:
            failed.append((ticker, str(e)))

        time.sleep(BATCH_PAUSE_SEC)

    if failed:
        print(f"\n   Failed tickers ({len(failed)}): {[t for t, _ in failed[:10]]}...")

    df = pd.DataFrame(rows)
    print(f"   Successfully enriched: {len(df)} companies above ${MIN_MARKET_CAP/1e6:.0f}M market cap")
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — Write to Excel screener
# ══════════════════════════════════════════════════════════════════════════════

def get_runway_flag(runway):
    if isinstance(runway, str):
        return "✅ CF+"
    if runway < 4:
        return "🔴 <4Q"
    elif runway < 8:
        return "🟡 4-8Q"
    else:
        return "🟢 >8Q"


def populate_screener(universe_df: pd.DataFrame, template_path: str, output_path: str):
    """
    Loads the existing screener template and populates it with the full universe.
    Preserves all formatting from the template.
    """
    print(f"\n[3/4] Populating screener with {len(universe_df)} companies...")

    wb = openpyxl.load_workbook(template_path)

    # ── Screener sheet ────────────────────────────────────────────────────────
    ws = wb["📊 Screener"]

    # Clear existing data rows (keep header at row 6)
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=7, max_row=max_row):
        for cell in row:
            cell.value = None

    # Colour helpers (matching original screener)
    def hex_fill(h): return PatternFill("solid", fgColor=h)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    TA_COLORS = {
        "Oncology":           ("FDE8E8", "C0392B"),
        "Rare Disease":       ("E8F0FE", "1A56DB"),
        "Immunology":         ("E8F5EE", "1A7A4A"),
        "CNS":                ("F3E8FD", "7B2FBE"),
        "Cardiometabolic":    ("FEF3E2", "C27803"),
    }
    STAGE_COLORS = {
        "Phase 1":    ("FDE8E8", "C0392B"),
        "Phase 1/2":  ("FEF3E2", "A05C00"),
        "Phase 2":    ("FEF3E2", "C27803"),
        "Phase 3":    ("E8F0FE", "1A56DB"),
        "Approved":   ("E8F5EE", "1A7A4A"),
        "Preclinical":("F5F5F5", "595959"),
    }
    MODALITY_COLORS = {
        "Small Molecule":     ("EAF4FB", "1A6FA8"),
        "Monoclonal Ab (mAb)":("FEF3E2", "A05C00"),
        "RNA":                ("F0FBF0", "1A7A4A"),
        "Gene Therapy":       ("F5E8FD", "6A1B9A"),
        "Gene Editing":       ("F5E8FD", "6A1B9A"),
        "Protein / Peptide":  ("FDE8F5", "8B1A6B"),
        "ADC / Bispecific":   ("FFF8E1", "795548"),
        "Cell Therapy":       ("FFF3E0", "E65100"),
    }

    # Sort by market cap descending
    universe_df = universe_df.sort_values("Market Cap ($M)", ascending=False).reset_index(drop=True)

    for row_idx, (_, co) in enumerate(universe_df.iterrows(), start=7):
        ws.row_dimensions[row_idx].height = 32
        fill = hex_fill("FFFFFF") if row_idx % 2 == 0 else hex_fill("F5F5F5")

        mkt_cap = co.get("Market Cap ($M)", 0) or 0
        cash    = co.get("Cash ($M)", 0) or 0
        burn    = co.get("Qtrly Burn ($M)", 0) or 0

        def w(col, val, fmt=None, font_kw=None, fill_override=None):
            c = ws.cell(row=row_idx, column=col)
            c.value = val
            c.font = Font(name="Arial", size=9, **(font_kw or {}))
            c.fill = fill_override or fill
            c.border = thin
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if fmt: c.number_format = fmt
            return c

        # Col 1: Ticker
        w(1, co["Ticker"], font_kw={"bold": True, "color": "1B2A4A"})
        # Col 2: Company Name
        c2 = w(2, co["Company Name"])
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        # Col 3: Exchange
        w(3, co.get("Exchange", ""))
        # Col 4: Therapy Area — left blank for manual/pipeline enrichment
        ta = co.get("Therapy Area", "")
        ta_bg, ta_fg = TA_COLORS.get(ta, ("F5F5F5", "595959"))
        w(4, ta,
          font_kw={"bold": bool(ta), "color": ta_fg},
          fill_override=hex_fill(ta_bg) if ta else fill)
        # Col 5-6: Sub-Indication, Modality — blank (pipeline enrichment needed)
        c5 = w(5, co.get("Sub_Indication", ""))
        c5.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        mod = co.get("Modality", "")
        mod_bg, mod_fg = MODALITY_COLORS.get(mod, ("F5F5F5", "595959"))
        w(6, mod,
          font_kw={"color": mod_fg},
          fill_override=hex_fill(mod_bg) if mod else fill)
        # Col 7: Lead Asset
        c7 = w(7, co.get("Lead_Asset", ""))
        c7.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        # Col 8: Dev Stage
        stage = co.get("Dev_Stage", "")
        st_bg, st_fg = STAGE_COLORS.get(stage, ("F5F5F5", "595959"))
        w(8, stage,
          font_kw={"bold": bool(stage), "color": st_fg},
          fill_override=hex_fill(st_bg) if stage else fill)
        # Col 9-11: Catalyst, Catalyst Type, Partners — blank
        w(9, co.get("Next_Catalyst", ""))
        c10 = w(10, co.get("Catalyst_Type", ""))
        c10.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c11 = w(11, co.get("Partners", ""))
        c11.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        # Col 12-14: Financials
        w(12, mkt_cap if mkt_cap else None, fmt="#,##0")
        w(13, cash if cash else None,    fmt="#,##0")
        w(14, burn if burn else None,    fmt="#,##0",
          font_kw={"color": "0000FF" if burn > 0 else "1A7A4A"})
        # Col 15: Runway formula
        r = row_idx
        rc = ws.cell(row=r, column=15)
        if burn and burn > 0:
            rc.value = f"=IFERROR(ROUND(M{r}/N{r},1),\"N/A\")"
        else:
            rc.value = "CF+" if cash > 0 else ""
        rc.font = Font(name="Arial", size=9)
        rc.fill = fill
        rc.border = thin
        rc.alignment = Alignment(horizontal="center", vertical="center")
        rc.number_format = "0.0"
        # Col 16: Runway flag formula
        fc = ws.cell(row=r, column=16)
        if burn and burn > 0:
            fc.value = (f'=IFERROR(IF(O{r}="CF+","✅ CF+",'
                        f'IF(O{r}<4,"🔴 <4Q",IF(O{r}<8,"🟡 4-8Q","🟢 >8Q"))),"—")')
        else:
            fc.value = "✅ CF+" if cash > 0 else ""
        fc.font = Font(name="Arial", size=9, bold=True)
        fc.fill = fill
        fc.border = thin
        fc.alignment = Alignment(horizontal="center", vertical="center")
        # Col 17: EV formula
        ec = ws.cell(row=r, column=17)
        ec.value = f"=IFERROR(L{r}-M{r},\"\")"
        ec.font = Font(name="Arial", size=9)
        ec.fill = fill
        ec.border = thin
        ec.alignment = Alignment(horizontal="center", vertical="center")
        ec.number_format = "#,##0"
        # Col 18-19: 52W High/Low
        w(18, co.get("52W High ($)") or None, fmt="#,##0.00")
        w(19, co.get("52W Low ($)") or None,  fmt="#,##0.00")
        # Col 20: Notes — prepopulate with business description snippet
        desc = co.get("Description", "")
        c20 = w(20, desc[:200] if desc else "")
        c20.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c20.font = Font(name="Arial", size=8, italic=True, color="595959")

    # Update autofilter range to cover new data
    last_row = 6 + len(universe_df)
    ws.auto_filter.ref = f"A6:{get_column_letter(20)}{last_row}"

    # Update title date
    ws["A4"].value = (
        f"Last Updated: {datetime.today().strftime('%d %b %Y')}     |     "
        f"Universe: {len(universe_df)} companies (SIC 2836, ≥$50M mkt cap)     |     "
        f"Data Sources: SEC EDGAR · yFinance"
    )

    # ── Financials sheet ──────────────────────────────────────────────────────
    ws_fin = wb["💰 Financials DB"]
    max_row_fin = ws_fin.max_row
    for row in ws_fin.iter_rows(min_row=6, max_row=max_row_fin):
        for cell in row:
            cell.value = None

    fin_fill_w = PatternFill("solid", fgColor="FFFFFF")
    fin_fill_g = PatternFill("solid", fgColor="E8F5EE")

    for row_idx, (_, co) in enumerate(universe_df.iterrows(), start=6):
        fill_f = fin_fill_w if row_idx % 2 == 0 else fin_fill_g
        ws_fin.row_dimensions[row_idx].height = 24

        def fw(col, val, fmt=None):
            c = ws_fin.cell(row=row_idx, column=col)
            c.value = val
            c.font = Font(name="Arial", size=9, bold=(col == 1))
            c.fill = fill_f
            c.border = thin
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt: c.number_format = fmt

        mkt  = co.get("Market Cap ($M)", 0) or 0
        cash = co.get("Cash ($M)", 0) or 0
        burn = co.get("Qtrly Burn ($M)", 0) or 0
        r    = row_idx

        fw(1,  co["Ticker"])
        c2f = ws_fin.cell(row=r, column=2)
        c2f.value = co["Company Name"]
        c2f.font = Font(name="Arial", size=9)
        c2f.fill = fill_f
        c2f.border = thin
        c2f.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        fw(3,  datetime.today().strftime("%b %Y"))
        fw(4,  mkt  if mkt  else None, "#,##0")
        fw(5,  co.get("Share Price ($)") or None, "#,##0.00")
        fw(6,  co.get("Shares Out (M)") or None, "#,##0")
        fw(7,  cash if cash else None, "#,##0")
        fw(8,  0,   "#,##0")   # Short-term investments — needs manual
        fw(9,  cash if cash else None, "#,##0")   # Total liquid = cash (simplified)
        fw(10, burn if burn else None, "#,##0")

        # Runway
        run_c = ws_fin.cell(row=r, column=11)
        if burn and burn > 0:
            run_c.value = f"=IFERROR(ROUND(G{r}/J{r},1),\"\")"
        else:
            run_c.value = "CF+"
        run_c.font = Font(name="Arial", size=9)
        run_c.fill = fill_f
        run_c.border = thin
        run_c.alignment = Alignment(horizontal="center", vertical="center")
        run_c.number_format = "0.0"

        # Flag
        flag_c = ws_fin.cell(row=r, column=12)
        if burn and burn > 0:
            flag_c.value = f'=IFERROR(IF(K{r}<4,"🔴 <4Q",IF(K{r}<8,"🟡 4-8Q","🟢 >8Q")),"—")'
        else:
            flag_c.value = "✅ CF+"
        flag_c.font = Font(name="Arial", size=9, bold=True)
        flag_c.fill = fill_f
        flag_c.border = thin
        flag_c.alignment = Alignment(horizontal="center", vertical="center")

        fw(13, "—")  # Revenue TTM — needs manual from 10-K
        fw(14, "—")  # Rev growth

        ev_c = ws_fin.cell(row=r, column=15)
        ev_c.value = f"=IFERROR(D{r}-G{r},\"\")"
        ev_c.font = Font(name="Arial", size=9)
        ev_c.fill = fill_f
        ev_c.border = thin
        ev_c.alignment = Alignment(horizontal="center", vertical="center")
        ev_c.number_format = "#,##0"

        src_c = ws_fin.cell(row=r, column=16)
        src_c.value = f"yFinance / SEC EDGAR, {datetime.today().strftime('%b %Y')}"
        src_c.font = Font(name="Arial", size=8, italic=True)
        src_c.fill = fill_f
        src_c.border = thin
        src_c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws_fin.auto_filter.ref = f"A5:{get_column_letter(16)}{5 + len(universe_df)}"

    wb.save(output_path)
    print(f"   Saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 4 — Save raw CSV for inspection / backup
# ══════════════════════════════════════════════════════════════════════════════

def save_raw_csv(df: pd.DataFrame, path: str):
    df.to_csv(path, index=False)
    print(f"\n[4/4] Raw universe saved to: {path}  ({len(df)} rows)")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 60)
    print("  BIOTECH UNIVERSE BUILDER — SIC 2836")
    print(f"  {datetime.today().strftime('%d %b %Y %H:%M')}")
    print("=" * 60)

    # Check template exists
    if not os.path.exists(INPUT_XLSX):
        print(f"\n❌ Template not found: {INPUT_XLSX}")
        print("   Place Biotech_Stock_Screener.xlsx in the same folder.")
        exit(1)

    # Step 1: Pull SIC 2836 universe from EDGAR
    sec_df = fetch_sec_sic_companies(SIC_CODE)

    if sec_df is None or len(sec_df) == 0:
        print("❌ No companies returned from EDGAR. Check your internet connection.")
        exit(1)

    tickers = sec_df["ticker"].dropna().str.upper().unique().tolist()
    print(f"\n   Ticker universe: {len(tickers)} unique tickers")

    # Step 2: Enrich with yFinance
    enriched_df = enrich_with_yfinance(tickers)

    if len(enriched_df) == 0:
        print("❌ yFinance returned no data. Check your connection.")
        exit(1)

    # Step 3: Merge SEC name/exchange with yFinance data
    final_df = enriched_df.copy()
    # Pipeline fields — blank by default, to be populated manually or via
    # the separate pipeline enrichment script (build_pipeline.py)
    for col in ["Therapy Area", "Sub_Indication", "Modality", "Lead_Asset",
                "Dev_Stage", "Next_Catalyst", "Catalyst_Type", "Partners"]:
        if col not in final_df.columns:
            final_df[col] = ""

    # Save raw CSV
    save_raw_csv(final_df, RAW_CSV)

    # Step 4: Populate screener
    populate_screener(final_df, INPUT_XLSX, OUTPUT_XLSX)

    print("\n✅ Done!")
    print(f"   Full screener: {OUTPUT_XLSX}")
    print(f"   Raw CSV backup: {RAW_CSV}")
    print(f"\n   Next steps:")
    print(f"   • Open {OUTPUT_XLSX} — financial columns are populated")
    print(f"   • Pipeline columns (Therapy Area, Modality, Stage, Catalyst)")
    print(f"     require manual entry or the build_pipeline.py enrichment script")
    print(f"   • Revenue TTM in Financials DB requires manual entry from 10-K")
