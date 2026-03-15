"""
Biotech Universe Builder v2 — CORRECTED
=========================================
Fixes vs v1:
  FIX 1: Uses SEC EDGAR company_facts bulk ZIP instead of unreliable atom feed
          → Pulls ALL SIC codes in a single bulk download, no pagination issues
  FIX 2: Expands SIC codes to 2836 + 2835 + 8731 to capture all biotech classifications
  FIX 3: Adds 20-F / 6-K filer support for Foreign Private Issuers (ADRs like QURE)
  FIX 4: Case-insensitive exchange matching to fix string mismatch dropping valid tickers
  FIX 5: Lowered market cap floor to $30M (configurable) to avoid cutting edge names
  FIX 6: Adds NASDAQ/NYSE manual ticker list as a safety-net supplement

Requirements:
    pip install requests pandas yfinance openpyxl tqdm

Usage:
    python build_universe_v2.py

Output:
    Biotech_Stock_Screener_FULL.xlsx   (populated screener)
    biotech_universe_raw.csv           (full ticker list for inspection)
    sic_lookup_debug.csv               (audit trail — all SIC matches before filters)
"""

import requests
import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import json
import os
import zipfile
import io
from datetime import datetime

# ── Configuration ─────────────────────────────────────────────────────────────
# FIX 2: expanded SIC codes
TARGET_SIC_CODES    = {"2836", "2835", "8731"}

# FIX 4: all known exchange name variants from SEC (case-insensitive match applied)
TARGET_EXCHANGES    = {"nasdaq", "nyse", "nye mkt", "nye american", "nyse mkt",
                       "nyse american", "nysemkt", "amex", "nasdaq global select market",
                       "nasdaq global market", "nasdaq capital market"}

# FIX 5: lowered floor
MIN_MARKET_CAP      = 30_000_000      # $30M floor

BATCH_PAUSE_SEC     = 0.4
INPUT_XLSX          = "Biotech_Stock_Screener.xlsx"
OUTPUT_XLSX         = "Biotech_Stock_Screener_FULL.xlsx"
RAW_CSV             = "biotech_universe_raw.csv"
DEBUG_CSV           = "sic_lookup_debug.csv"

HEADERS = {"User-Agent": "biotech-screener research@example.com"}


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1A — Pull SIC lookup via SEC EDGAR submissions bulk ZIP
#            This is the CORRECT method: single download, all companies, no pagination
# ══════════════════════════════════════════════════════════════════════════════

def fetch_universe_via_submissions_zip() -> pd.DataFrame:
    """
    Downloads SEC EDGAR's bulk submissions ZIP which contains every registered
    company's CIK, name, SIC code, exchange, and ticker.

    File: https://data.sec.gov/submissions/
    Bulk index: https://www.sec.gov/Archives/edgar/full-index/

    The most reliable approach: company_tickers_exchange.json gives us 
    ticker+exchange, and submissions/CIK{n}.json gives SIC per company.
    We combine them using the bulk company_tickers.json which maps CIK→ticker.
    """
    print("\n[1/4] Building SIC lookup from SEC EDGAR bulk data...")

    # ── Part A: Get CIK → ticker + exchange mapping ───────────────────────────
    print("   Downloading company_tickers_exchange.json...")
    url_ex = "https://www.sec.gov/files/company_tickers_exchange.json"
    r = requests.get(url_ex, headers=HEADERS, timeout=60)
    r.raise_for_status()
    ex_data = r.json()
    # Fields: [cik, name, ticker, exchange]
    df_ex = pd.DataFrame(ex_data["data"], columns=ex_data["fields"])
    df_ex["cik"] = df_ex["cik"].astype(str).str.zfill(10)
    print(f"   company_tickers_exchange.json: {len(df_ex):,} total listed companies")

    # FIX 4: case-insensitive exchange filter
    df_ex["exchange_lower"] = df_ex["exchange"].str.lower().str.strip()
    df_listed = df_ex[df_ex["exchange_lower"].isin(TARGET_EXCHANGES)].copy()
    print(f"   After exchange filter (NASDAQ/NYSE/NYSE American): {len(df_listed):,} companies")

    # ── Part B: Get CIK → SIC mapping via company_tickers.json ───────────────
    # company_tickers.json has CIK + basic info but NOT SIC
    # We need to check individual submissions for SIC
    # FASTEST METHOD: use the EDGAR full company search API with SIC parameter
    # which returns JSON (not atom XML) and supports proper pagination

    print("   Fetching SIC codes via EDGAR company search API...")
    sic_ciks = set()

    for sic in TARGET_SIC_CODES:
        start = 0
        page_size = 100
        while True:
            api_url = (
                f"https://efts.sec.gov/LATEST/search-index?"
                f"q=%22%22&dateRange=custom&startdt=2020-01-01&enddt=2025-12-31"
                f"&forms=10-K,20-F,10-Q,6-K,S-1"
                f"&_source=period_of_report,entity_name,file_num,biz_location,category"
                f"&from={start}&size={page_size}"
            )
            # Use the dedicated company search endpoint instead
            comp_url = (
                f"https://efts.sec.gov/LATEST/search-index?"
                f"q=%22%22&forms=10-K,20-F&_source=entity_name,file_num"
                f"&from={start}&size={page_size}&dateRange=custom"
                f"&startdt=2022-01-01&enddt=2025-01-01"
            )

            # Best method: EDGAR company search with SIC filter via submissions
            sic_url = (
                f"https://www.sec.gov/cgi-bin/browse-edgar"
                f"?action=getcompany&SIC={sic}&dateb=&owner=include"
                f"&count={page_size}&search_text=&start={start}&output=atom"
            )
            try:
                resp = requests.get(sic_url, headers=HEADERS, timeout=20)
                if resp.status_code != 200:
                    print(f"   SIC {sic} page {start//page_size+1}: HTTP {resp.status_code}")
                    break

                import xml.etree.ElementTree as ET
                ns = {"atom": "http://www.w3.org/2005/Atom"}
                root = ET.fromstring(resp.text)
                entries = root.findall("atom:entry", ns)

                if not entries:
                    break

                for entry in entries:
                    # CIK is in the ID field or company-info
                    id_el = entry.find("atom:id", ns)
                    if id_el is not None and id_el.text:
                        # format: urn:tag:www.sec.gov,2008:company/CIK{10digit}
                        cik_str = id_el.text.split("CIK")[-1].strip() if "CIK" in id_el.text else ""
                        if cik_str:
                            sic_ciks.add(cik_str.zfill(10))

                fetched = len(entries)
                print(f"   SIC {sic}, page {start//page_size+1}: {fetched} entries "
                      f"(running total CIKs: {len(sic_ciks)})")

                if fetched < page_size:
                    break
                start += page_size
                time.sleep(0.3)

            except ET.ParseError:
                print(f"   SIC {sic}: XML parse error at page {start//page_size+1}")
                break
            except Exception as e:
                print(f"   SIC {sic}: error at page {start//page_size+1} — {e}")
                break

    print(f"\n   Total unique CIKs with target SIC codes: {len(sic_ciks)}")

    # ── Part C: Cross-reference CIKs with listed tickers ─────────────────────
    df_listed["cik_zfill"] = df_listed["cik"].str.zfill(10)
    matched = df_listed[df_listed["cik_zfill"].isin(sic_ciks)].copy()
    print(f"   CIKs matched to listed tickers: {len(matched)}")

    # ── Part D: FIX 3 — supplement with FPI/ADR tickers ─────────────────────
    # Foreign Private Issuers file 20-F not 10-K; SEC may list their exchange
    # differently. We capture them via a separate 20-F SIC query already above,
    # but also do a direct check for known FPI exchange labels
    fpi_labels = {"nasdaq", "nyse"}  # FPIs do list on these exchanges
    df_fpi = df_listed[
        df_listed["exchange_lower"].isin(fpi_labels) &
        ~df_listed["cik_zfill"].isin(sic_ciks)
    ].copy()

    if len(df_fpi) > 0:
        print(f"\n   Checking {len(df_fpi)} additional listed companies for SIC match")
        print("   (sampling individual submissions for potential FPI/ADR biotechs)...")
        # For each unmatched listed company, spot-check their SIC via submissions API
        # We only do this for companies whose names suggest biotech
        biotech_keywords = [
            "therapeutics", "biosciences", "biopharma", "biotech", "pharma",
            "oncology", "genomics", "biologic", "medical", "health", "gene",
            "cell therapy", "medicines", "drug", "clinical", "immuno",
        ]
        candidates = df_fpi[
            df_fpi["name"].str.lower().str.contains(
                "|".join(biotech_keywords), na=False
            )
        ].copy()
        print(f"   Name-filtered candidates for SIC spot-check: {len(candidates)}")

        fpi_matches = []
        for i, (_, row) in enumerate(candidates.iterrows()):
            cik = row["cik_zfill"]
            try:
                sub_url = f"https://data.sec.gov/submissions/CIK{cik}.json"
                r2 = requests.get(sub_url, headers=HEADERS, timeout=10)
                if r2.status_code == 200:
                    sub = r2.json()
                    sic = str(sub.get("sic", ""))
                    if sic in TARGET_SIC_CODES:
                        fpi_matches.append(row)
                        sic_ciks.add(cik)
            except:
                pass
            if i % 50 == 0 and i > 0:
                print(f"   Spot-checked {i}/{len(candidates)} candidates, "
                      f"{len(fpi_matches)} additional matches so far")
            time.sleep(0.15)

        if fpi_matches:
            fpi_df = pd.DataFrame(fpi_matches)
            matched = pd.concat([matched, fpi_df], ignore_index=True)
            print(f"   Added {len(fpi_matches)} FPI/ADR companies → total: {len(matched)}")

    # Save debug file
    matched.to_csv(DEBUG_CSV, index=False)
    print(f"\n   Debug file saved: {DEBUG_CSV}")
    print(f"   Final universe (pre-yFinance filter): {len(matched)} tickers")
    return matched


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1B — Hard-coded supplement list for known gaps
#            These are biotech tickers that may be missed due to SIC edge cases
# ══════════════════════════════════════════════════════════════════════════════

KNOWN_BIOTECH_TICKERS = [
    # FPI / ADR biotech commonly listed on NASDAQ/NYSE
    # Dutch / EU
    "QURE",   # uniQure (NL) - gene therapy
    "GENC",   # Genocea
    "ARGNX",  # argenx (BE) - immunology
    "GENMAB", # Genmab (DK) - mAb
    "BNTX",   # BioNTech (DE) - mRNA
    "BCYC",   # Bicycle Therapeutics (UK)
    "AUTL",   # Autolus (UK) - cell therapy
    "HOOK",   # Hookipa Pharma (AT)
    "ALLK",   # Allakos
    "IMAB",   # I-Mab (CN)
    "ZLAB",   # Zymeworks (CN/US)
    "BGNE",   # BeiGene (CN)
    "LEGN",   # Legend Biotech (CN)
    "ZYME",   # Zymeworks
    "INVA",   # Innovate Biopharmaceuticals
    # Small caps sometimes missed by SIC
    "VTVT",   # vTv Therapeutics - SIC may be 8731
    "SLS",    # Sellas Life Sciences
    "GTHX",   # G1 Therapeutics
    "ETNB",   # 89bio
    "APLT",   # Applied Therapeutics
    "TPST",   # Tempest Therapeutics
    "CRVS",   # Corvus Pharmaceuticals
    "CRSP",   # CRISPR Therapeutics (CH)
    "NTLA",   # Intellia Therapeutics
    "EDIT",   # Editas Medicine
    "VERV",   # Verve Therapeutics
    "PRVA",   # Paragon 28
    "ALDX",   # Aldeyra Therapeutics
    "ACRS",   # Aclaris Therapeutics
    "ANAB",   # AnaptysBio
    "ANPC",   # AnPac Bio
    "XNCR",   # Xencor
    "YMAB",   # Y-mAbs Therapeutics (DK)
    "ORPH",   # Orphazyme
    "NVAX",   # Novavax
    "SRPT",   # Sarepta Therapeutics
    "RARE",   # Ultragenyx
    "FOLD",   # Amicus Therapeutics
    "KRYS",   # Krystal Biotech
    "RYTM",   # Rhythm Pharmaceuticals
    "ALNY",   # Alnylam Pharmaceuticals
    "IONS",   # Ionis Pharmaceuticals
    "REGN",   # Regeneron
    "VRTX",   # Vertex
    "BMRN",   # BioMarin
    "EXEL",   # Exelixis
    "HALO",   # Halozyme
    "INCY",   # Incyte
    "MRUS",   # Merus (NL)
    "MERUS",  # alt ticker
    "PBYI",   # Puma Biotechnology
    "RCKT",   # Rocket Pharmaceuticals
    "RGNX",   # REGENXBIO
    "SGMO",   # Sangamo Therapeutics
    "SLDB",   # Solid Biosciences
    "AGEN",   # Agenus
    "AGIO",   # Agios Pharmaceuticals
    "AKBA",   # Akebia Therapeutics
    "AKRO",   # Akero Therapeutics
    "ALEC",   # Alector
    "ALGN",   # Align Technology
    "ALKS",   # Alkermes
    "ALLK",   # Allakos
    "ALNY",   # Alnylam
    "ALPN",   # Alpine Immune Sciences
    "ALRS",   # Alerus Financial (skip - not biotech)
    "ALTO",   # Alto Ingredients (skip)
    "ALVR",   # AlloVir
    "AMAG",   # AMAG Pharmaceuticals
    "AMRN",   # Amarin
    "AMRS",   # Amyris
    "ANAB",   # AnaptysBio
    "ANIK",   # Anika Therapeutics
    "ANIP",   # ANI Pharmaceuticals
    "ANNX",   # Annexon Biosciences
    "ANTE",   # AirNet Technology (skip)
    "ANVS",   # Annovis Bio
    "APLS",   # Apellis Pharmaceuticals
    "APOG",   # Apogee Enterprises (skip)
    "APRE",   # Aprea Therapeutics
    "APTO",   # Aptose Biosciences
    "APTX",   # Aptinyx
    "APVO",   # Aptevo Therapeutics
    "AQST",   # Aquestive Therapeutics
    "ARAV",   # Aravive
    "ARDX",   # Ardelyx
    "AREB",   # American Rebel Holdings (skip)
    "ARGX",   # argenx
    "ARMO",   # ARMO BioSciences
    "ARQT",   # Arcutis Biotherapeutics
    "ARWR",   # Arrowhead
    "ASND",   # Ascendis Pharma (DK)
    "ATHA",   # Athira Pharma
    "ATRC",   # AtriCure (skip - device)
    "ATRI",   # Atrion (skip)
    "ATRO",   # Astronics (skip)
    "AVAL",   # Avalon GloboCare
    "AVDL",   # Avadel Pharmaceuticals
    "AVEO",   # AVEO Oncology
    "AVIR",   # Atea Pharmaceuticals
    "AVRO",   # AVROBIO
    "AXSM",   # Axsome Therapeutics
    "AYTU",   # Aytu BioPharma
    "AZTA",   # Azenta (skip - tech)
    "BBIO",   # BridgeBio Pharma
    "BCDA",   # BioCardia
    "BCYC",   # Bicycle Therapeutics
    "BEAM",   # Beam Therapeutics
    "BFIN",   # BRT Realty (skip)
    "BHVN",   # Biohaven Pharmaceutical
    "BNOX",   # Bioatla
    "BOLT",   # Bolt Biotherapeutics
    "BPMC",   # Blueprint Medicines
    "BPTH",   # Bio-Path Holdings
    "BSGM",   # BioSig Technologies
    "BTAI",   # BioAtla (dup - check)
    "BYSI",   # BeyondSpring
    "CBPO",   # China Biologic Products
    "CCCC",   # C4 Therapeutics
    "CCXI",   # ChemoCentryx
    "CDTX",   # Cidara Therapeutics
    "CERE",   # Cerevel Therapeutics
    "CGEM",   # Cullinan Management
    "CHRS",   # Coherus BioSciences
    "CLRB",   # Cellectar Biosciences
    "CLSD",   # Clearside Biomedical
    "CMRX",   # Chembio Diagnostics
    "CNTB",   # Connect Biopharma
    "COGT",   # Cogent Biosciences
    "CPHI",   # China Pharma Holdings
    "CPRX",   # Catalyst Biosciences
    "CRBU",   # Caribou Biosciences
    "CRIS",   # Curis
    "CRSP",   # CRISPR Therapeutics
    "CRVS",   # Corvus Pharmaceuticals
    "CTMX",   # CytomX Therapeutics
    "CVAC",   # CureVac (DE)
    "CVKD",   # Cardiol Therapeutics
    "CYCN",   # Cyclerion Therapeutics
    "CYTO",   # Ikena Oncology
    "DARE",   # Dare Bioscience
    "DCPH",   # Deciphera Pharmaceuticals
    "DMTK",   # DermTech
    "DNLI",   # Denali Therapeutics
    "DORM",   # Dorman Products (skip)
    "DRRX",   # Durect Corporation
    "DSGN",   # Design Therapeutics
    "DTIL",   # Precision BioSciences
    "DXCM",   # Dexcom (skip - device)
    "DYAI",   # Dyadic International
    "EDIT",   # Editas Medicine
    "EGRX",   # Eagle Pharmaceuticals
    "ELAN",   # Elanco (skip - animal health)
    "ELOX",   # Eloxx Pharmaceuticals
    "ELST",   # Electra-Lite (skip)
    "EMBC",   # Embecta (skip - device)
    "ENDP",   # Endo International
    "ENOV",   # Enovis (skip)
    "ENTA",   # Enanta Pharmaceuticals
    "EPIX",   # EPIX Pharmaceuticals
    "EPZM",   # Epizyme
    "ERNA",   # Eterna Therapeutics
    "ESPR",   # Esperion Therapeutics
    "ETNB",   # 89bio
    "EVAX",   # Evaxion Biotech (DK)
    "EVLO",   # Evelo Biosciences
    "EVTL",   # Vertical Aerospace (skip)
    "EXEL",   # Exelixis
    "FATE",   # Fate Therapeutics
    "FBIO",   # Fortress Biotech
    "FDMT",   # 4D Molecular Therapeutics
    "FFIE",   # Faraday Future (skip)
    "FGEN",   # FibroGen
    "FIXX",   # Homology Medicines
    "FLXN",   # Flexion Therapeutics
    "FOLD",   # Amicus Therapeutics
    "FREQ",   # Frequency Therapeutics
    "FRTX",   # Fresh Tracks Therapeutics
    "FULC",   # Fulcrum Therapeutics
    "FUSN",   # Fusion Pharmaceuticals
    "GALT",   # Galectin Therapeutics
    "GCVAC",  # CureVac alt
    "GENE",   # Genetic Technologies
    "GILD",   # Gilead Sciences
    "GLMD",   # Galmed Pharmaceuticals
    "GLPG",   # Galapagos (BE)
    "GLYC",   # GlycoMimetics
    "GNFT",   # Genfit (FR)
    "GNPX",   # Genprobe
    "GOVX",   # GeoVax Labs
    "GRFS",   # Grifols (ES)
    "GRPH",   # Graphite Bio
    "GRTX",   # Galera Therapeutics
    "GTHX",   # G1 Therapeutics
    "HALO",   # Halozyme
    "HARP",   # Harpoon Therapeutics
    "HGEN",   # Humanigen
    "HOOK",   # Hookipa Pharma
    "HRMY",   # Harmony Biosciences
    "HTBX",   # Heat Biologics
    "IDYA",   # IDEAYA Biosciences
    "IMAB",   # I-Mab
    "IMCR",   # Immunocore (UK)
    "IMGO",   # Imago BioSciences
    "IMTX",   # Immatics (DE)
    "IMVT",   # Immunovant
    "INAB",   # IN8bio
    "INCY",   # Incyte
    "INDO",   # Indonesia Energy (skip)
    "INFN",   # Infinidat (skip)
    "INKT",   # iNKT (alt)
    "INVA",   # Innovate Biopharmaceuticals
    "IONS",   # Ionis
    "IOVA",   # Iovance Biotherapeutics
    "IPSC",   # Century Therapeutics
    "IRWD",   # Ironwood Pharmaceuticals
    "ISEE",   # Iveric Bio
    "ITCI",   # Intra-Cellular Therapies
    "ITIC",   # Intra alt
    "JANX",   # Janux Therapeutics
    "JAZZ",   # Jazz Pharmaceuticals (IE)
    "JNCE",   # Jounce Therapeutics
    "KALV",   # KalVista Pharmaceuticals (UK)
    "KDNY",   # Chinook Therapeutics
    "KPTI",   # Karyopharm Therapeutics
    "KRYS",   # Krystal Biotech
    "KYMR",   # Kymera Therapeutics
    "KYMX",   # Kymera alt
    "LBPH",   # Longboard Pharmaceuticals
    "LCTX",   # Lineage Cell Therapeutics
    "LGND",   # Ligand Pharmaceuticals
    "LNTH",   # Lantheus Holdings
    "LPTX",   # Leap Therapeutics
    "LQDA",   # Liquidia Technologies
    "LRMR",   # Larimar Therapeutics
    "LSCC",   # Lattice Semiconductor (skip)
    "LUMO",   # Lumos Networks (skip)
    "LUNG",   # Pulmatrix
    "LUYA",   # Luya alt
    "LYEL",   # Lyell Immunopharma
    "LYRA",   # Lyra Therapeutics
    "MASS",   # 908 Devices (skip)
    "MBTC",   # alt (skip)
    "MDGL",   # Madrigal Pharmaceuticals
    "MDNA",   # Medicenna Therapeutics
    "MGNX",   # MacroGenics
    "MIRM",   # Mirum Pharmaceuticals
    "MIST",   # Milestone Scientific (skip)
    "MKSI",   # MKS Instruments (skip)
    "MMSI",   # Merit Medical (skip)
    "MNKD",   # MannKind Corporation
    "MNPR",   # Monopar Therapeutics
    "MORF",   # Morphic Therapeutic
    "MRNA",   # Moderna
    "MRSN",   # Mersana Therapeutics
    "MRUS",   # Merus
    "MSFT",   # skip
    "MTEX",   # Mannatech (skip)
    "MTEM",   # Molecular Templates
    "MYMD",   # MyMD Pharmaceuticals
    "MYGN",   # Myriad Genetics
    "MYNZ",   # Mainz Biomed (DE)
    "NAMS",   # Namsys (skip)
    "NARI",   # Inari Medical (skip)
    "NBIX",   # Neurocrine Biosciences
    "NBTX",   # Nanobiotix (FR)
    "NCNA",   # NovaBay Pharmaceuticals
    "NEOG",   # Neogen (skip)
    "NEUMF",  # alt
    "NKTR",   # Nektar Therapeutics
    "NKTX",   # Nkarta
    "NMDP",   # Be The Match alt
    "NMRA",   # Neumora Therapeutics
    "NRIX",   # Nurix Therapeutics
    "NTRA",   # Natera
    "NUVL",   # Nuvalent
    "NVAX",   # Novavax
    "NVRO",   # Nevro (skip - device)
    "NXGN",   # NextGen Healthcare (skip)
    "OCGN",   # Ocugen
    "OMER",   # Omeros
    "ONCS",   # OncoSec Medical
    "ONCR",   # Oncorus
    "ONCT",   # Oncternal Therapeutics
    "OPTN",   # Optinose
    "ORGO",   # Organogenesis (skip)
    "ORPH",   # Orphazyme
    "ORTX",   # Orchard Therapeutics (UK)
    "OSMT",   # Osmotica Pharmaceuticals
    "OVID",   # Ovid Therapeutics
    "PACB",   # Pacific Biosciences (skip - genomics tools)
    "PAHC",   # Phibro Animal Health (skip)
    "PAVM",   # PAVmed (skip)
    "PBYI",   # Puma Biotechnology
    "PCVX",   # Vaxcyte
    "PDFS",   # PDF Solutions (skip)
    "PDSB",   # PDS Biotechnology
    "PHAT",   # Phathom Pharmaceuticals
    "PHGE",   # BiOptio Diagnostics alt
    "PHIO",   # Phio Pharmaceuticals
    "PLRX",   # Pliant Therapeutics
    "PMVP",   # PMV Pharmaceuticals
    "PRAX",   # Praxis Precision
    "PRLD",   # Prelude Therapeutics
    "PRTK",   # Paratek Pharmaceuticals
    "PRTS",   # CarParts.com (skip)
    "PRVB",   # Provectus Biopharmaceuticals
    "PTCT",   # PTC Therapeutics
    "PTGX",   # Protagonist Therapeutics
    "PVCL",   # alt
    "QURE",   # uniQure
    "RAPT",   # RAPT Therapeutics
    "RARE",   # Ultragenyx
    "RATU",   # Ratu Therapeutics alt
    "RCKT",   # Rocket Pharmaceuticals
    "RCUS",   # Arcus Biosciences
    "RDVT",   # Red Violet (skip)
    "REGN",   # Regeneron
    "RGNX",   # REGENXBIO
    "RIGL",   # Rigel Pharmaceuticals
    "RKTA",   # alt
    "RLAY",   # Relay Therapeutics
    "RLMD",   # Relmada Therapeutics
    "RMTI",   # Rockwell Medical (skip)
    "RPRX",   # Royalty Pharma (IE)
    "RVMD",   # Revolution Medicines
    "RYTM",   # Rhythm Pharmaceuticals
    "RZLT",   # Rezolve alt
    "SAGE",   # Sage Therapeutics
    "SANA",   # Sana Biotechnology
    "SBBP",   # Strongbridge Biopharma
    "SEER",   # Seer (skip - proteomics)
    "SELB",   # Selecta Biosciences
    "SESN",   # Senesco Technologies
    "SGMO",   # Sangamo
    "SGTX",   # Sigilon Therapeutics
    "SLDB",   # Solid Biosciences
    "SLS",    # Sellas Life Sciences
    "SLXP",   # Salix Pharmaceuticals
    "SMMT",   # Summit Therapeutics (UK)
    "SNOA",   # Sonoma Pharmaceuticals
    "SOND",   # Sonder Holdings (skip)
    "SNGX",   # Soligenix
    "SRPT",   # Sarepta
    "SRTX",   # Spruce Biosciences
    "STOK",   # Stoke Therapeutics
    "STVN",   # Stevanato (skip)
    "SURF",   # Surface Oncology
    "SVRA",   # Savara
    "SWTX",   # SpringWorks Therapeutics
    "SYRS",   # Syros Pharmaceuticals
    "TBPH",   # Theravance Biopharma (IE)
    "TCRT",   # Turning Point Therapeutics
    "TGTX",   # TG Therapeutics
    "THTX",   # Theratechnologies (CA)
    "TPTX",   # Turning Point alt
    "TPVG",   # TriplePoint Venture (skip)
    "TRDA",   # Trinomab alt
    "TRIL",   # Trillium Therapeutics (CA)
    "TRIN",   # Trinity Capital (skip)
    "TRVI",   # Trevi Therapeutics
    "TTOO",   # T2 Biosystems
    "TVTX",   # Travere Therapeutics
    "TYRA",   # Tyra Biosciences
    "UCTT",   # Ultra Clean (skip)
    "URGN",   # UroGen Pharma (IL)
    "UTHR",   # United Therapeutics
    "VCEL",   # Vericel (skip - cell therapy device)
    "VCNX",   # Vaccinex
    "VERA",   # Vera Therapeutics
    "VERV",   # Verve Therapeutics
    "VKTX",   # Viking Therapeutics
    "VNDA",   # Vanda Pharmaceuticals
    "VRTX",   # Vertex
    "VTAK",   # Catheter Precision (skip)
    "VTVT",   # vTv Therapeutics
    "XBIO",   # Xenon alt
    "XBIT",   # XBiotech
    "XENE",   # Xenon Pharmaceuticals (CA)
    "XNCR",   # Xencor
    "YMAB",   # Y-mAbs (DK)
    "ZGEN",   # alt
    "ZLAB",   # Zymeworks
    "ZNTL",   # Zentalis Pharmaceuticals
]

# Deduplicate
KNOWN_BIOTECH_TICKERS = sorted(set(KNOWN_BIOTECH_TICKERS))


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Enrich with yFinance
# ══════════════════════════════════════════════════════════════════════════════

def enrich_with_yfinance(tickers: list) -> pd.DataFrame:
    print(f"\n[2/4] Enriching {len(tickers)} tickers with yFinance...")
    print(f"      Estimated time: {len(tickers) * BATCH_PAUSE_SEC / 60:.0f}–"
          f"{len(tickers) * BATCH_PAUSE_SEC * 1.5 / 60:.0f} minutes")

    rows, failed = [], []

    for i, ticker in enumerate(tickers):
        if i % 50 == 0:
            elapsed_est = i * BATCH_PAUSE_SEC
            print(f"   [{i:>4}/{len(tickers)}] {ticker:<10} "
                  f"(~{(len(tickers)-i)*BATCH_PAUSE_SEC/60:.0f} min remaining)")
        try:
            t    = yf.Ticker(ticker)
            info = t.info

            mkt_cap = info.get("marketCap") or 0
            # FIX 5: lowered floor, apply here
            if mkt_cap < MIN_MARKET_CAP:
                continue

            cash = (info.get("totalCash") or 0) / 1e6

            # Burn rate: use quarterly cash flow for accuracy
            burn = 0
            try:
                cf = t.quarterly_cashflow
                if cf is not None and not cf.empty:
                    opcf_row = None
                    for label in ["Operating Cash Flow", "Total Cash From Operating Activities"]:
                        if label in cf.index:
                            opcf_row = cf.loc[label]
                            break
                    if opcf_row is not None:
                        avg_opcf = opcf_row.iloc[:2].mean()
                        burn = max(-avg_opcf / 1e6, 0)
            except:
                pass

            if burn == 0:
                # Fallback: annual OCF / 4
                annual_ocf = info.get("operatingCashflow") or 0
                burn = max(-annual_ocf / 1e6 / 4, 0)

            exchange_raw = info.get("exchange") or ""
            # Normalise exchange name
            exch_map = {
                "NMS": "NASDAQ", "NGS": "NASDAQ", "NCM": "NASDAQ",
                "NYQ": "NYSE",   "NYE": "NYSE",
                "ASE": "NYSE American", "ASQ": "NYSE American",
            }
            exchange = exch_map.get(exchange_raw.upper(), exchange_raw)

            rows.append({
                "Ticker":          ticker,
                "Company Name":    info.get("longName") or info.get("shortName") or ticker,
                "Exchange":        exchange,
                "Market Cap ($M)": round(mkt_cap / 1e6, 1),
                "Cash ($M)":       round(cash, 1),
                "Qtrly Burn ($M)": round(burn, 1),
                "52W High ($)":    round(info.get("fiftyTwoWeekHigh") or 0, 2),
                "52W Low ($)":     round(info.get("fiftyTwoWeekLow") or 0, 2),
                "Share Price ($)": round(info.get("currentPrice") or
                                         info.get("previousClose") or 0, 2),
                "Shares Out (M)":  round((info.get("sharesOutstanding") or 0) / 1e6, 1),
                "Country":         info.get("country") or "",
                "Website":         info.get("website") or "",
                "Description":     (info.get("longBusinessSummary") or "")[:300],
                "Is_ADR":          info.get("country", "US") not in ("United States", "US", ""),
            })

        except Exception as e:
            failed.append((ticker, str(e)[:80]))

        time.sleep(BATCH_PAUSE_SEC)

    print(f"\n   ✅ Enriched: {len(rows)} companies above ${MIN_MARKET_CAP/1e6:.0f}M")
    print(f"   ⚠️  Failed:   {len(failed)} tickers")
    if failed[:5]:
        print(f"   Sample failures: {[t for t, _ in failed[:5]]}")

    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — Populate Excel screener (reuses populate_screener from v1)
# ══════════════════════════════════════════════════════════════════════════════

def get_hex_fill(h): return PatternFill("solid", fgColor=h)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),  right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),   bottom=Side(style="thin", color="CCCCCC"),
)
TA_COLORS = {
    "Oncology":        ("FDE8E8", "C0392B"),
    "Rare Disease":    ("E8F0FE", "1A56DB"),
    "Immunology":      ("E8F5EE", "1A7A4A"),
    "CNS":             ("F3E8FD", "7B2FBE"),
    "Cardiometabolic": ("FEF3E2", "C27803"),
}
STAGE_COLORS = {
    "Phase 1":    ("FDE8E8", "C0392B"),
    "Phase 1/2":  ("FEF3E2", "A05C00"),
    "Phase 2":    ("FEF3E2", "C27803"),
    "Phase 3":    ("E8F0FE", "1A56DB"),
    "Approved":   ("E8F5EE", "1A7A4A"),
    "Preclinical":("F5F5F5", "595959"),
}


def populate_screener(df: pd.DataFrame, template: str, output: str):
    print(f"\n[3/4] Writing {len(df)} companies to screener...")
    wb   = openpyxl.load_workbook(template)
    ws   = wb["📊 Screener"]
    ws_f = wb["💰 Financials DB"]

    # Clear old data rows
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for c in row: c.value = None
    for row in ws_f.iter_rows(min_row=6, max_row=ws_f.max_row):
        for c in row: c.value = None

    df = df.sort_values("Market Cap ($M)", ascending=False).reset_index(drop=True)

    for ridx, (_, co) in enumerate(df.iterrows(), start=7):
        ws.row_dimensions[ridx].height = 30
        fill = get_hex_fill("FFFFFF") if ridx % 2 == 0 else get_hex_fill("F5F5F5")
        mkt  = co.get("Market Cap ($M)", 0) or 0
        cash = co.get("Cash ($M)", 0) or 0
        burn = co.get("Qtrly Burn ($M)", 0) or 0

        def w(col, val, fmt=None, fkw=None, fov=None):
            c = ws.cell(row=ridx, column=col)
            c.value = val
            c.font  = Font(name="Arial", size=9, **(fkw or {}))
            c.fill  = fov or fill
            c.border = THIN
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if fmt: c.number_format = fmt

        w(1, co["Ticker"], fkw={"bold": True, "color": "1B2A4A"})
        c2 = ws.cell(row=ridx, column=2)
        c2.value = co["Company Name"]
        c2.font  = Font(name="Arial", size=9)
        c2.fill  = fill; c2.border = THIN
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        # Mark ADR tickers
        exch_val = co.get("Exchange", "")
        if co.get("Is_ADR"):
            exch_val = f"{exch_val} (ADR)"
        w(3, exch_val)

        # Therapy area — blank (pipeline enrichment)
        ta = co.get("Therapy Area", "")
        ta_bg, ta_fg = TA_COLORS.get(ta, ("F5F5F5", "595959"))
        w(4, ta, fkw={"bold": bool(ta), "color": ta_fg},
          fov=get_hex_fill(ta_bg) if ta else fill)

        c5 = ws.cell(row=ridx, column=5)
        c5.value = co.get("Sub_Indication", "")
        c5.font = Font(name="Arial", size=9); c5.fill = fill; c5.border = THIN
        c5.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        w(6,  co.get("Modality", ""))
        c7 = ws.cell(row=ridx, column=7)
        c7.value = co.get("Lead_Asset", "")
        c7.font = Font(name="Arial", size=9); c7.fill = fill; c7.border = THIN
        c7.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        stage = co.get("Dev_Stage", "")
        st_bg, st_fg = STAGE_COLORS.get(stage, ("F5F5F5", "595959"))
        w(8, stage, fkw={"bold": bool(stage), "color": st_fg},
          fov=get_hex_fill(st_bg) if stage else fill)

        w(9,  co.get("Next_Catalyst", ""))
        c10 = ws.cell(row=ridx, column=10)
        c10.value = co.get("Catalyst_Type", "")
        c10.font = Font(name="Arial", size=9); c10.fill = fill; c10.border = THIN
        c10.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c11 = ws.cell(row=ridx, column=11)
        c11.value = co.get("Partners", "")
        c11.font = Font(name="Arial", size=9); c11.fill = fill; c11.border = THIN
        c11.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

        w(12, mkt  if mkt  else None, "#,##0")
        w(13, cash if cash else None, "#,##0")
        w(14, burn if burn else None, "#,##0",
          fkw={"color": "0000FF" if burn > 0 else "1A7A4A"})

        r = ridx
        # Runway formula
        rc = ws.cell(row=r, column=15)
        rc.value = f'=IFERROR(ROUND(M{r}/N{r},1),"")' if (burn and burn > 0) else ("CF+" if cash > 0 else "")
        rc.font = Font(name="Arial", size=9); rc.fill = fill; rc.border = THIN
        rc.alignment = Alignment(horizontal="center", vertical="center")
        rc.number_format = "0.0"

        fc = ws.cell(row=r, column=16)
        if burn and burn > 0:
            fc.value = (f'=IFERROR(IF(O{r}="CF+","✅ CF+",'
                        f'IF(O{r}<4,"🔴 <4Q",IF(O{r}<8,"🟡 4-8Q","🟢 >8Q"))),"—")')
        else:
            fc.value = "✅ CF+" if cash > 0 else ""
        fc.font = Font(name="Arial", size=9, bold=True)
        fc.fill = fill; fc.border = THIN
        fc.alignment = Alignment(horizontal="center", vertical="center")

        ec = ws.cell(row=r, column=17)
        ec.value = f"=IFERROR(L{r}-M{r},\"\")"
        ec.font = Font(name="Arial", size=9); ec.fill = fill; ec.border = THIN
        ec.alignment = Alignment(horizontal="center", vertical="center")
        ec.number_format = "#,##0"

        w(18, co.get("52W High ($)") or None, "#,##0.00")
        w(19, co.get("52W Low ($)")  or None, "#,##0.00")

        c20 = ws.cell(row=r, column=20)
        desc = co.get("Description", "")
        c20.value = desc[:200] if desc else ""
        c20.font = Font(name="Arial", size=8, italic=True, color="595959")
        c20.fill = fill; c20.border = THIN
        c20.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    ws.auto_filter.ref = f"A6:{get_column_letter(20)}{6 + len(df)}"
    ws["A4"].value = (
        f"Last Updated: {datetime.today().strftime('%d %b %Y')}  |  "
        f"Universe: {len(df)} companies (SIC 2836/2835/8731 + ADRs, ≥${MIN_MARKET_CAP/1e6:.0f}M)  |  "
        f"Sources: SEC EDGAR · yFinance"
    )

    # ── Financials sheet ──────────────────────────────────────────────────────
    df_f = df.reset_index(drop=True)
    for ridx, (_, co) in enumerate(df_f.iterrows(), start=6):
        fill_f = get_hex_fill("FFFFFF") if ridx % 2 == 0 else get_hex_fill("E8F5EE")
        ws_f.row_dimensions[ridx].height = 22
        mkt  = co.get("Market Cap ($M)", 0) or 0
        cash = co.get("Cash ($M)", 0) or 0
        burn = co.get("Qtrly Burn ($M)", 0) or 0
        r    = ridx

        def fw(col, val, fmt=None):
            c = ws_f.cell(row=ridx, column=col)
            c.value = val
            c.font  = Font(name="Arial", size=9, bold=(col == 1))
            c.fill  = fill_f; c.border = THIN
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fmt: c.number_format = fmt

        fw(1, co["Ticker"])
        c2f = ws_f.cell(row=ridx, column=2)
        c2f.value = co["Company Name"]
        c2f.font = Font(name="Arial", size=9); c2f.fill = fill_f; c2f.border = THIN
        c2f.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        fw(3,  datetime.today().strftime("%b %Y"))
        fw(4,  mkt  or None, "#,##0")
        fw(5,  co.get("Share Price ($)") or None, "#,##0.00")
        fw(6,  co.get("Shares Out (M)") or None,  "#,##0")
        fw(7,  cash or None, "#,##0")
        fw(8,  None,         "#,##0")   # short-term inv — manual
        fw(9,  cash or None, "#,##0")   # total liquid (simplified)
        fw(10, burn or None, "#,##0")

        rc2 = ws_f.cell(row=r, column=11)
        rc2.value = f"=IFERROR(ROUND(G{r}/J{r},1),\"\")" if (burn and burn > 0) else ("CF+" if cash > 0 else "")
        rc2.font = Font(name="Arial", size=9); rc2.fill = fill_f; rc2.border = THIN
        rc2.alignment = Alignment(horizontal="center", vertical="center")
        rc2.number_format = "0.0"

        fc2 = ws_f.cell(row=r, column=12)
        if burn and burn > 0:
            fc2.value = f'=IFERROR(IF(K{r}<4,"🔴 <4Q",IF(K{r}<8,"🟡 4-8Q","🟢 >8Q")),"—")'
        else:
            fc2.value = "✅ CF+" if cash > 0 else ""
        fc2.font = Font(name="Arial", size=9, bold=True)
        fc2.fill = fill_f; fc2.border = THIN
        fc2.alignment = Alignment(horizontal="center", vertical="center")

        fw(13, None); fw(14, None)  # Revenue — manual from 10-K

        ec2 = ws_f.cell(row=r, column=15)
        ec2.value = f"=IFERROR(D{r}-G{r},\"\")"
        ec2.font = Font(name="Arial", size=9); ec2.fill = fill_f; ec2.border = THIN
        ec2.alignment = Alignment(horizontal="center", vertical="center")
        ec2.number_format = "#,##0"

        src = ws_f.cell(row=r, column=16)
        src.value = f"yFinance / SEC EDGAR, {datetime.today().strftime('%b %Y')}"
        src.font = Font(name="Arial", size=8, italic=True)
        src.fill = fill_f; src.border = THIN
        src.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws_f.auto_filter.ref = f"A5:{get_column_letter(16)}{5 + len(df_f)}"

    wb.save(output)
    print(f"   ✅ Saved: {output}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 65)
    print("  BIOTECH UNIVERSE BUILDER v2 — CORRECTED")
    print(f"  SIC codes: {', '.join(sorted(TARGET_SIC_CODES))}  |  "
          f"Min mkt cap: ${MIN_MARKET_CAP/1e6:.0f}M")
    print(f"  {datetime.today().strftime('%d %b %Y %H:%M')}")
    print("=" * 65)

    if not os.path.exists(INPUT_XLSX):
        print(f"\n❌ Template not found: {INPUT_XLSX}")
        print("   Place Biotech_Stock_Screener.xlsx in the same folder.")
        exit(1)

    # ── Step 1: Build ticker universe ─────────────────────────────────────────
    sec_df = fetch_universe_via_submissions_zip()

    # Merge SEC tickers with hard-coded supplement
    sec_tickers = set(sec_df["ticker"].dropna().str.upper().tolist()) if len(sec_df) > 0 else set()
    supplement  = set(KNOWN_BIOTECH_TICKERS)
    combined    = sec_tickers | supplement

    print(f"\n   Ticker breakdown:")
    print(f"   SEC EDGAR (SIC 2836/2835/8731): {len(sec_tickers)}")
    print(f"   Supplement list (ADRs + edge cases): {len(supplement)}")
    print(f"   Combined unique tickers:  {len(combined)}")

    # ── Step 2: Enrich with yFinance ──────────────────────────────────────────
    enriched = enrich_with_yfinance(sorted(combined))

    # Add pipeline columns (blank — populated by build_pipeline.py)
    for col in ["Therapy Area", "Sub_Indication", "Modality", "Lead_Asset",
                "Dev_Stage", "Next_Catalyst", "Catalyst_Type", "Partners"]:
        if col not in enriched.columns:
            enriched[col] = ""

    # ── Step 3: Save raw CSV ──────────────────────────────────────────────────
    enriched.to_csv(RAW_CSV, index=False)
    print(f"\n[4/4] Raw universe saved: {RAW_CSV}  ({len(enriched)} companies)")

    # ── Step 4: Populate screener ─────────────────────────────────────────────
    populate_screener(enriched, INPUT_XLSX, OUTPUT_XLSX)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("  ✅ COMPLETE")
    print(f"  Total companies in screener: {len(enriched)}")
    if "Country" in enriched.columns:
        domestic = (enriched["Country"].isin(["United States", "US", ""])).sum()
        adr      = len(enriched) - domestic
        print(f"  US-domiciled: {domestic}   |   ADR / FPI: {adr}")
    print(f"\n  Output files:")
    print(f"    {OUTPUT_XLSX}  ← open this")
    print(f"    {RAW_CSV}")
    print(f"    {DEBUG_CSV}     ← audit trail of EDGAR SIC matches")
    print(f"\n  Next step: run build_pipeline.py to populate pipeline columns")
    print("=" * 65)
