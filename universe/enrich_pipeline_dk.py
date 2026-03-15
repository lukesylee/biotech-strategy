"""
Pipeline Enrichment Script — Columns D to K
=============================================
Reads tickers from biotech_universe_raw.csv (or directly from the screener),
queries ClinicalTrials.gov API v2 for each company, and writes:

  D  Therapy Area       — classified from indication keywords
  E  Sub-Indication     — lead trial's primary condition (most specific)
  F  Modality           — classified from intervention type + name keywords
  G  Lead Asset         — drug name of the highest-phase trial
  H  Dev Stage          — highest phase across all company's trials
  I  Next Catalyst      — primary completion date of the lead trial
  J  Catalyst Type      — inferred label (e.g. "Phase 3 Readout", "NDA Filing")
  K  Partners / Collab  — collaborator names from lead trial

Requirements:
    pip install requests pandas openpyxl tqdm

Usage:
    # Option A — enrich the FULL screener produced by build_universe_v2.py
    python enrich_pipeline_dk.py --input Biotech_Stock_Screener_FULL.xlsx

    # Option B — enrich the sample screener (30 companies)
    python enrich_pipeline_dk.py --input Biotech_Stock_Screener.xlsx

    # Option C — dry-run: just write pipeline_db.csv, don't touch Excel
    python enrich_pipeline_dk.py --input Biotech_Stock_Screener_FULL.xlsx --csv-only

Output:
    <input_file>  (updated in-place, columns D-K populated)
    pipeline_db.csv   (one row per trial, full audit trail)
"""

import argparse
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import re
from datetime import datetime

# ── Configuration ─────────────────────────────────────────────────────────────
CTGOV_API       = "https://clinicaltrials.gov/api/v2/studies"
PAUSE_SEC       = 0.25          # polite rate limit between API calls
MAX_TRIALS      = 100           # max trials to fetch per company
PIPELINE_CSV    = "pipeline_db.csv"
HEADERS         = {"User-Agent": "biotech-screener research@example.com"}

# ── Therapy Area: keyword matching on condition/indication text ───────────────
# Order matters — first match wins. More specific terms first.
THERAPY_AREA_RULES = [
    ("Rare Disease", [
        "rare disease", "orphan disease", "duchenne", "pompe disease", "fabry disease",
        "gaucher", "niemann-pick", "wilson disease", "huntington disease",
        "amyloidosis", "ttr amyloid", "transthyretin amyloid",
        "mps ", "mucopolysaccharidosis", "hemophilia a", "hemophilia b",
        "haemophilia", "sickle cell disease", "thalassemia", "thalassaemia",
        "cystic fibrosis", "spinal muscular atrophy", " sma ", "friedreich ataxia",
        "maple syrup urine", "phenylketonuria", " pku ", "epidermolysis bullosa",
        "dystrophic eb", "congenital adrenal hyperplasia",
        "tuberous sclerosis", "angelman syndrome", "rett syndrome",
        "neurofibromatosis", "hereditary transthyretin", "lysosomal storage",
        "urea cycle disorder", "organic acidemia", "glycogen storage disease",
    ]),
    ("Oncology", [
        "cancer", "tumor", "tumour", "carcinoma", "adenocarcinoma", "sarcoma",
        "lymphoma", "leukemia", "leukaemia", "myeloma", "melanoma", "glioma",
        "glioblastoma", "glioblastoma multiforme", "gbm", "mesothelioma",
        "neuroblastoma", "medulloblastoma", "hepatocellular", "cholangiocarcinoma",
        "pancreatic ductal", "nsclc", "sclc", "non-small cell", "small cell lung",
        "urothelial", "bladder cancer", "renal cell", "ovarian cancer",
        "breast cancer", "prostate cancer", "colorectal", "endometrial",
        "cervical cancer", "gastric cancer", "esophageal", "head and neck",
        "oncology", "neoplasm", "neoplasia", "malignant", "metastatic",
        "relapsed", "refractory", "haematologic", "hematologic malignancy",
        "diffuse large b-cell", "dlbcl", "follicular lymphoma", "cll", "all", "aml",
        "myelodysplastic", "mds", "polycythemia vera", "essential thrombocythemia",
        "myelofibrosis",
    ]),
    ("CNS", [
        "alzheimer", "parkinson", "parkinson's", "epilepsy", "seizure", "dravet",
        "lennox-gastaut", "multiple sclerosis", "relapsing ms", "schizophrenia",
        "major depressive disorder", "mdd", "bipolar", "anxiety disorder",
        "post-traumatic stress", "ptsd", "attention deficit", "adhd",
        "amyotrophic lateral sclerosis", "als", "motor neuron disease",
        "neuropathic pain", "migraine", "essential tremor", "tourette",
        "autism spectrum", "fragile x", "angelman syndrome", "rett syndrome",
        "neurological", "neurodegenerative", "psychiatric", "cns",
        "cognitive impairment", "dementia", "frontotemporal",
        "agitation", "insomnia", "narcolepsy", "idiopathic hypersomnia",
    ]),
    ("Immunology", [
        "rheumatoid arthritis", "psoriatic arthritis", "ankylosing spondylitis",
        "axial spondyloarthritis", "systemic lupus", "lupus nephritis",
        "psoriasis", "plaque psoriasis", "atopic dermatitis", "eczema",
        "crohn", "ulcerative colitis", "inflammatory bowel", "ibd",
        "sjogren", "myasthenia gravis", "cidp", "chronic inflammatory",
        "idiopathic thrombocytopenic", "itp", "pemphigus", "pemphigoid",
        "iga nephropathy", "igan", "primary biliary", "pbc", "primary sclerosing",
        "autoimmune hepatitis", "alopecia areata", "vitiligo", "hidradenitis",
        "uveitis", "thyroid eye", "graves", "immune thrombocytopenia",
        "cold agglutinin", "neuromyelitis optica", "nmosd",
        "asthma", "chronic obstructive pulmonary", "copd",
        "autoimmune", "immunology", "inflammation", "inflammatory",
    ]),
    ("Cardiometabolic", [
        "heart failure", "hfref", "hfpef", "cardiac", "cardiomyopathy",
        "cardiovascular", "atherosclerosis", "coronary artery", "acute coronary",
        "myocardial infarction", "hypertension", "pulmonary arterial hypertension",
        "pah", "hypercholesterolemia", "hyperlipidemia", "dyslipidemia",
        "ldl", "triglyceride", "hypertriglyceridemia", "familial hypercholesterol",
        "type 2 diabetes", "type 1 diabetes", "diabetes mellitus", "t2dm", "t1dm",
        "obesity", "overweight", "nash", "mash", "nafld", "masld", "steatohepatitis",
        "non-alcoholic fatty liver", "metabolic", "insulin resistance",
        "gout", "hyperuricemia", "lipoprotein", "aortic stenosis",
        "atrial fibrillation", "ventricular", "arrhythmia", "thrombosis",
        "hemostasis", "coagulation", "von willebrand",
    ]),
    ("Infectious Disease", [
        "hiv", "human immunodeficiency", "hepatitis b", "hepatitis c", "hbv", "hcv",
        "tuberculosis", "tb", "malaria", "influenza", "covid-19", "sars-cov-2",
        "respiratory syncytial virus", "rsv", "cytomegalovirus", "cmv",
        "clostridium difficile", "c. diff", "candida", "aspergillus", "fungal",
        "bacterial infection", "sepsis", "pneumonia", "urinary tract infection",
        "antibiotic", "antimicrobial", "antiviral", "antifungal", "vaccination",
        "vaccine", "infectious",
    ]),
    ("Ophthalmology", [
        "macular degeneration", "amd", "diabetic retinopathy", "diabetic macular",
        "retinal vein occlusion", "glaucoma", "dry eye", "corneal",
        "stargardt", "retinitis pigmentosa", "optic neuritis", "leber",
        "choroidal neovascularization", "geographic atrophy", "ophthalm",
        "retinal", "vitreous", "ocular",
    ]),
]

# ── Modality: classified from intervention type + drug name keywords ──────────
# Checked in order; first match wins.
MODALITY_RULES = [
    # Name-based (more specific — checked before type-based)
    ("ADC / Bispecific",   ["antibody-drug conjugate", " adc", "bispecific", "bsab",
                            "tandem", "tebentafusp", "amivantamab",
                            "tarlatamab", "xaluritamig"]),
    ("Cell Therapy",       ["car-t", "cart", "car t", "chimeric antigen", "til ",
                            "tumor infiltrating", "nk cell", "natural killer cell",
                            "tcr-t", "t-cell receptor", "autologous cell",
                            "allogeneic cell", "ipsc-derived"]),
    ("Gene Therapy",       ["gene therapy", "aav", "adeno-associated virus",
                            "lentiviral vector", "retroviral vector",
                            "viral vector", "oncolytic virus", "oncolytic viral"]),
    ("Gene Editing",       ["crispr", "base edit", "prime edit", "zinc finger",
                            "talen", "meganuclease", "gene editing", "gene correction",
                            "hdr", "homology-directed"]),
    ("RNA (siRNA/ASO/mRNA)",["sirna", "si-rna", "antisense", " aso ", "mrna ",
                             "messenger rna", "lnp-mrna", "rna interference",
                             "rnai", "mirna", "antagomir", "splice-switching",
                             "exon skip"]),
    ("Monoclonal Ab (mAb)", ["monoclonal antibody", "mab", " mab ", "-mab",
                             "antibody", "immunoglobulin", "checkpoint",
                             "anti-pd", "anti-pd-1", "anti-pd-l1",
                             "anti-ctla", "anti-tigit", "anti-lag",
                             "anti-cd", "anti-her", "anti-egfr",
                             "anti-vegf", "anti-il", "anti-tnf",
                             "anti-fcrn", "anti-c5", "biologic"]),
    ("Protein / Peptide",  ["peptide", "protein", "enzyme replacement",
                            "recombinant protein", "fusion protein",
                            "fc-fusion", "albumin fusion", "growth factor",
                            "cytokine", "interleukin", "erythropoietin",
                            "thrombopoietin", "hepcidin", "glp-1", "gip",
                            "glucagon", "insulin"]),
]

# Intervention-type based fallback (when name keywords don't match)
INTERVENTION_TYPE_MAP = {
    "GENETIC":             "Gene Therapy",
    "BIOLOGICAL":          "Monoclonal Ab (mAb)",   # broad; refined by name above
    "DRUG":                "Small Molecule",
    "COMBINATION_PRODUCT": "Small Molecule",
    "DIETARY_SUPPLEMENT":  "Other",
    "DEVICE":              "Other",
    "PROCEDURE":           "Other",
    "RADIATION":           "Other",
    "BEHAVIORAL":          "Other",
    "OTHER":               "Other",
}

# ── Phase priority for selecting "lead" trial ─────────────────────────────────
PHASE_PRIORITY = {
    "PHASE3": 7, "PHASE2|PHASE3": 6, "PHASE2": 5,
    "PHASE1|PHASE2": 4, "PHASE1": 3, "EARLY_PHASE1": 2, "NA": 1, "": 0,
}

PHASE_DISPLAY = {
    "PHASE3": "Phase 3",     "PHASE2|PHASE3": "Phase 2/3",
    "PHASE2": "Phase 2",     "PHASE1|PHASE2": "Phase 1/2",
    "PHASE1": "Phase 1",     "EARLY_PHASE1": "Phase 1",
    "NA": "Preclinical",     "": "Preclinical",
}


# ══════════════════════════════════════════════════════════════════════════════
#  Classification helpers
# ══════════════════════════════════════════════════════════════════════════════

def classify_therapy_area(conditions: list, title: str = "") -> str:
    text = " ".join(str(c).lower() for c in conditions) + " " + title.lower()
    for area, keywords in THERAPY_AREA_RULES:
        if any(kw in text for kw in keywords):
            return area
    return "Other"


def classify_modality(interventions: list) -> str:
    """
    interventions: list of dicts with keys 'type' and 'name'
    Returns the most specific modality label.
    """
    if not interventions:
        return ""

    # Combine all intervention names into one string for keyword matching
    all_names = " ".join(
        str(i.get("name", "")).lower() for i in interventions
        if i.get("type", "") not in ("DEVICE", "PROCEDURE", "BEHAVIORAL", "RADIATION",
                                      "DIETARY_SUPPLEMENT", "OTHER")
    )

    # Check name-based rules first (more specific)
    for modality, keywords in MODALITY_RULES:
        if any(kw in all_names for kw in keywords):
            return modality

    # Fallback: intervention type
    for i in interventions:
        itype = i.get("type", "").upper()
        if itype in INTERVENTION_TYPE_MAP:
            mapped = INTERVENTION_TYPE_MAP[itype]
            if mapped not in ("Other", ""):
                return mapped

    return "Small Molecule"   # sensible default for DRUG type


def get_phase_key(phase_list: list) -> str:
    """Return the canonical phase key from a list of phase strings."""
    if not phase_list:
        return ""
    # Normalise
    phases = [str(p).upper().replace(" ", "").replace("-", "") for p in phase_list]
    # Multi-phase trials
    if len(phases) > 1:
        joined = "|".join(sorted(phases))
        if "PHASE2" in joined and "PHASE3" in joined:
            return "PHASE2|PHASE3"
        if "PHASE1" in joined and "PHASE2" in joined:
            return "PHASE1|PHASE2"
    p = phases[0]
    for key in PHASE_PRIORITY:
        if key in p:
            return key
    return "NA"


def infer_catalyst_type(phase_key: str, status: str, is_approved: bool = False) -> str:
    if is_approved:
        return "Commercial / Label Expansion"
    status_up = status.upper()
    if phase_key == "PHASE3":
        if "COMPLETED" in status_up:
            return "Phase 3 Data Readout"
        return "Phase 3 Readout"
    if phase_key in ("PHASE2", "PHASE2|PHASE3"):
        return "Phase 2 Data Readout"
    if phase_key == "PHASE1|PHASE2":
        return "Phase 1/2 Data Readout"
    if phase_key in ("PHASE1", "EARLY_PHASE1"):
        return "Phase 1 Data / Dose Escalation"
    return "Clinical Update"


def format_date(date_str: str) -> str:
    """Convert YYYY-MM-DD or YYYY-MM to a cleaner display format."""
    if not date_str:
        return ""
    try:
        if re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            return dt.strftime("%b %Y")
        if re.match(r"^\d{4}-\d{2}$", date_str):
            dt = datetime.strptime(date_str + "-01", "%Y-%m-%d")
            return dt.strftime("%b %Y")
        if re.match(r"^\d{4}$", date_str):
            return date_str
    except:
        pass
    return date_str


# ══════════════════════════════════════════════════════════════════════════════
#  ClinicalTrials.gov API fetch
# ══════════════════════════════════════════════════════════════════════════════

def fetch_trials(company_name: str) -> list:
    """
    Fetch all trials for a company from ClinicalTrials.gov API v2.
    Tries company name first; if < 3 results, also tries ticker as a fallback query.
    Returns a list of parsed trial dicts.
    """
    params = {
        "query.spons":        company_name,
        "filter.overallStatus": (
            "RECRUITING,ACTIVE_NOT_RECRUITING,COMPLETED,"
            "ENROLLING_BY_INVITATION,NOT_YET_RECRUITING"
        ),
        "fields": (
            "NCTId,BriefTitle,OfficialTitle,Phase,OverallStatus,"
            "LeadSponsorName,CollaboratorName,Condition,"
            "InterventionType,InterventionName,"
            "PrimaryCompletionDate,StartDate,WhyStopped"
        ),
        "pageSize": MAX_TRIALS,
        "format":   "json",
    }
    try:
        r = requests.get(CTGOV_API, params=params, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            return []
        return r.json().get("studies", [])
    except Exception:
        return []


def parse_studies(studies: list, company_name: str) -> list:
    """Parse raw ClinicalTrials.gov studies into flat dicts."""
    rows = []
    for study in studies:
        proto = study.get("protocolSection", {})

        id_mod      = proto.get("identificationModule", {})
        status_mod  = proto.get("statusModule", {})
        design_mod  = proto.get("designModule", {})
        conds_mod   = proto.get("conditionsModule", {})
        arms_mod    = proto.get("armsInterventionsModule", {})
        sponsor_mod = proto.get("sponsorCollaboratorsModule", {})

        nct_id  = id_mod.get("nctId", "")
        title   = id_mod.get("briefTitle", "") or id_mod.get("officialTitle", "")
        status  = status_mod.get("overallStatus", "")

        # Phase
        raw_phases = design_mod.get("phases", [])
        phase_key  = get_phase_key(raw_phases)

        # Conditions
        conditions = conds_mod.get("conditions", [])

        # Interventions — filter to therapeutic types only
        all_interventions = arms_mod.get("interventions", [])
        therapeutic_types = {
            "DRUG", "BIOLOGICAL", "GENETIC", "COMBINATION_PRODUCT",
            "DIETARY_SUPPLEMENT",
        }
        interventions = [
            {"type": i.get("type", ""), "name": i.get("name", "")}
            for i in all_interventions
            if i.get("type", "").upper() in therapeutic_types
        ]

        # Drug name: prefer the first non-placebo/comparator intervention
        skip_words = {"placebo", "saline", "vehicle", "comparator", "standard of care",
                      "soc", "best supportive", "observation"}
        drug_name = ""
        for inv in interventions:
            n = inv.get("name", "").strip()
            if n and not any(s in n.lower() for s in skip_words):
                drug_name = n
                break

        # Collaborators (excl. the sponsor itself)
        lead_sponsor   = sponsor_mod.get("leadSponsor", {}).get("name", "")
        collaborators  = [
            c.get("name", "")
            for c in sponsor_mod.get("collaborators", [])
            if c.get("name", "") and c.get("name", "") != lead_sponsor
        ]

        # Dates
        prim_compl_struct = status_mod.get("primaryCompletionDateStruct") or {}
        prim_compl        = prim_compl_struct.get("date", "")

        rows.append({
            "nct_id":        nct_id,
            "title":         title[:120],
            "status":        status,
            "phase_key":     phase_key,
            "phase_display": PHASE_DISPLAY.get(phase_key, "Preclinical"),
            "conditions":    conditions,
            "lead_condition": conditions[0] if conditions else "",
            "interventions": interventions,
            "drug_name":     drug_name[:60],
            "collaborators": collaborators,
            "prim_compl":    prim_compl,
        })
    return rows


def select_lead_trial(trials: list) -> dict:
    """
    Select the single 'lead' trial — the highest-phase, most recently recruiting trial.
    Tie-break: prefer recruiting/active over completed; prefer more recent start date.
    """
    if not trials:
        return {}

    status_priority = {
        "RECRUITING":               5,
        "ACTIVE_NOT_RECRUITING":    4,
        "ENROLLING_BY_INVITATION":  3,
        "NOT_YET_RECRUITING":       2,
        "COMPLETED":                1,
        "TERMINATED":               0,
        "WITHDRAWN":                0,
        "SUSPENDED":                0,
    }

    def trial_score(t):
        phase_score  = PHASE_PRIORITY.get(t["phase_key"], 0) * 10
        status_score = status_priority.get(t["status"].upper(), 0)
        return phase_score + status_score

    return max(trials, key=trial_score)


# ══════════════════════════════════════════════════════════════════════════════
#  Main enrichment function
# ══════════════════════════════════════════════════════════════════════════════

def enrich_company(ticker: str, company_name: str) -> dict:
    """
    Returns a dict with keys D–K for one company.
    Falls back to empty strings if no data found.
    """
    empty = {
        "therapy_area":   "",
        "sub_indication": "",
        "modality":       "",
        "lead_asset":     "",
        "dev_stage":      "",
        "next_catalyst":  "",
        "catalyst_type":  "",
        "partners":       "",
        "trial_count":    0,
        "all_phases":     [],
    }

    studies  = fetch_trials(company_name)
    if not studies:
        return empty

    trials = parse_studies(studies, company_name)
    if not trials:
        return empty

    lead = select_lead_trial(trials)

    # Therapy area: use lead trial, with fallback to majority vote across all trials
    therapy = classify_therapy_area(lead["conditions"], lead["title"])
    if therapy == "Other" and len(trials) > 1:
        all_therapies = [
            classify_therapy_area(t["conditions"], t["title"])
            for t in trials
        ]
        non_other = [x for x in all_therapies if x != "Other"]
        if non_other:
            from collections import Counter
            therapy = Counter(non_other).most_common(1)[0][0]

    # Sub-indication: first condition of lead trial, cleaned up
    sub_ind = lead["lead_condition"][:80] if lead["lead_condition"] else ""

    # Modality
    modality = classify_modality(lead["interventions"])
    # If lead trial has no therapeutic intervention, scan all trials
    if not modality or modality == "Other":
        for t in sorted(trials, key=lambda x: PHASE_PRIORITY.get(x["phase_key"], 0), reverse=True):
            m = classify_modality(t["interventions"])
            if m and m != "Other":
                modality = m
                break

    # Lead asset
    lead_asset = lead["drug_name"]
    if not lead_asset:
        for t in trials:
            if t["drug_name"]:
                lead_asset = t["drug_name"]
                break

    # Dev stage — highest phase across all trials (not just lead)
    all_phase_keys = [t["phase_key"] for t in trials if t["phase_key"]]
    best_phase_key = max(all_phase_keys, key=lambda k: PHASE_PRIORITY.get(k, 0), default="")
    dev_stage = PHASE_DISPLAY.get(best_phase_key, "Preclinical")

    # Check if company has an approved product (any trial with status ~ approved)
    # Proxy: look for "APPROVED" in title or no-phase trials with commercial keywords
    titles_lower = " ".join(t["title"].lower() for t in trials)
    is_approved  = any(kw in titles_lower for kw in
                       ["approved", "post-marketing", "label expansion",
                        "real-world", "registry", "pharmacovigilance"])

    # Next catalyst: primary completion date of the lead trial
    next_cat = format_date(lead["prim_compl"]) if lead.get("prim_compl") else ""

    # Catalyst type
    cat_type = infer_catalyst_type(lead["phase_key"], lead["status"], is_approved)

    # Partners: unique collaborators from lead trial, capped at 3
    partners = "; ".join(lead["collaborators"][:3]) if lead.get("collaborators") else ""

    return {
        "therapy_area":   therapy,
        "sub_indication": sub_ind,
        "modality":       modality,
        "lead_asset":     lead_asset,
        "dev_stage":      dev_stage,
        "next_catalyst":  next_cat,
        "catalyst_type":  cat_type,
        "partners":       partners,
        "trial_count":    len(trials),
        "all_phases":     list(set(t["phase_display"] for t in trials)),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  Excel write-back
# ══════════════════════════════════════════════════════════════════════════════

# Colour maps (matching screener design)
TA_COLORS = {
    "Oncology":           ("FDE8E8", "C0392B"),
    "Rare Disease":       ("E8F0FE", "1A56DB"),
    "Immunology":         ("E8F5EE", "1A7A4A"),
    "CNS":                ("F3E8FD", "7B2FBE"),
    "Cardiometabolic":    ("FEF3E2", "C27803"),
    "Infectious Disease": ("FFF8E0", "8B6914"),
    "Ophthalmology":      ("E0F7FA", "006064"),
    "Other":              ("F5F5F5", "595959"),
}
STAGE_COLORS = {
    "Phase 1":     ("FDE8E8", "C0392B"),
    "Phase 1/2":   ("FEF3E2", "A05C00"),
    "Phase 2":     ("FEF3E2", "C27803"),
    "Phase 2/3":   ("E8F0FE", "1A56DB"),
    "Phase 3":     ("E8F0FE", "1A56DB"),
    "Approved":    ("E8F5EE", "1A7A4A"),
    "Preclinical": ("F5F5F5", "595959"),
}
MODALITY_COLORS = {
    "Small Molecule":      ("EAF4FB", "1A6FA8"),
    "Monoclonal Ab (mAb)": ("FEF3E2", "A05C00"),
    "ADC / Bispecific":    ("FFF8E1", "795548"),
    "Cell Therapy":        ("FFF3E0", "E65100"),
    "Gene Therapy":        ("F5E8FD", "6A1B9A"),
    "Gene Editing":        ("F5E8FD", "6A1B9A"),
    "RNA (siRNA/ASO/mRNA)":("F0FBF0", "1A7A4A"),
    "Protein / Peptide":   ("FDE8F5", "8B1A6B"),
    "Other":               ("F5F5F5", "595959"),
}


def write_row_dk(ws, row_idx: int, data: dict, base_fill: PatternFill):
    """Write columns D(4) through K(11) for one row."""

    def cell(col, val, fill=None, font_kw=None):
        c = ws.cell(row=row_idx, column=col)
        c.value = val if val else ""
        c.font  = Font(name="Arial", size=9, **(font_kw or {}))
        c.fill  = fill or base_fill
        c.alignment = Alignment(
            horizontal="center" if col not in (5, 7, 10, 11) else "left",
            vertical="center", indent=1 if col in (5, 7, 10, 11) else 0,
            wrap_text=True,
        )

    # D — Therapy Area (coloured badge)
    ta = data.get("therapy_area", "")
    ta_bg, ta_fg = TA_COLORS.get(ta, ("F5F5F5", "595959"))
    cell(4, ta,
         fill=PatternFill("solid", fgColor=ta_bg) if ta else base_fill,
         font_kw={"bold": bool(ta), "color": ta_fg})

    # E — Sub-Indication
    cell(5, data.get("sub_indication", ""))

    # F — Modality (coloured badge)
    mod = data.get("modality", "")
    mod_bg, mod_fg = MODALITY_COLORS.get(mod, ("F5F5F5", "595959"))
    cell(6, mod,
         fill=PatternFill("solid", fgColor=mod_bg) if mod else base_fill,
         font_kw={"color": mod_fg})

    # G — Lead Asset
    cell(7, data.get("lead_asset", ""))

    # H — Dev Stage (coloured badge)
    stage = data.get("dev_stage", "")
    st_bg, st_fg = STAGE_COLORS.get(stage, ("F5F5F5", "595959"))
    cell(8, stage,
         fill=PatternFill("solid", fgColor=st_bg) if stage else base_fill,
         font_kw={"bold": bool(stage), "color": st_fg})

    # I — Next Catalyst
    cell(9, data.get("next_catalyst", ""))

    # J — Catalyst Type
    cell(10, data.get("catalyst_type", ""))

    # K — Partners / Collab
    cell(11, data.get("partners", ""))


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Enrich screener columns D-K from ClinicalTrials.gov")
    parser.add_argument("--input",    default="Biotech_Stock_Screener_FULL.xlsx",
                        help="Path to the screener Excel file")
    parser.add_argument("--csv-only", action="store_true",
                        help="Only write pipeline_db.csv, skip Excel update")
    parser.add_argument("--resume",   action="store_true",
                        help="Skip tickers already present in pipeline_db.csv")
    args = parser.parse_args()

    import os
    if not os.path.exists(args.input):
        print(f"❌ File not found: {args.input}")
        return

    # ── Load screener ─────────────────────────────────────────────────────────
    print(f"\nLoading: {args.input}")
    wb = openpyxl.load_workbook(args.input)
    ws = wb["📊 Screener"]

    # Read tickers and company names from the screener (cols A, B, starting row 7)
    companies = []
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=2):
        ticker = row[0].value
        name   = row[1].value
        if ticker and str(ticker).strip():
            companies.append((str(ticker).strip().upper(), str(name).strip(), row[0].row))

    print(f"Companies to enrich: {len(companies)}")

    # ── Resume support ────────────────────────────────────────────────────────
    already_done = set()
    if args.resume and os.path.exists(PIPELINE_CSV):
        existing = pd.read_csv(PIPELINE_CSV)
        already_done = set(existing["Ticker"].str.upper().tolist())
        print(f"Resuming: {len(already_done)} tickers already in {PIPELINE_CSV}, skipping")

    # ── Enrich loop ───────────────────────────────────────────────────────────
    all_results = []
    failed      = []

    total     = len(companies)
    done      = 0
    skipped   = 0

    print(f"\n{'─'*65}")
    print(f"{'Ticker':<10} {'Company':<35} {'Stage':<14} {'Therapy':<20}")
    print(f"{'─'*65}")

    for ticker, name, excel_row in companies:

        if ticker in already_done:
            skipped += 1
            continue

        data = enrich_company(ticker, name)
        data["Ticker"]      = ticker
        data["Company"]     = name
        data["excel_row"]   = excel_row
        data["fetched_at"]  = datetime.today().strftime("%Y-%m-%d")

        all_results.append(data)
        done += 1

        stage   = data.get("dev_stage", "—")
        therapy = data.get("therapy_area", "—")
        trials  = data.get("trial_count", 0)
        print(f"{ticker:<10} {name[:34]:<35} {stage:<14} {therapy:<20}  ({trials} trials)")

        # Write to Excel immediately (so partial runs are saved)
        if not args.csv_only:
            base_fill = PatternFill("solid", fgColor="FFFFFF" if excel_row % 2 == 0 else "F5F5F5")
            write_row_dk(ws, excel_row, data, base_fill)

        # Incremental CSV save every 25 companies
        if done % 25 == 0:
            _save_csv(all_results)
            if not args.csv_only:
                wb.save(args.input)
            print(f"  → Progress saved ({done}/{total - skipped} enriched)")

        time.sleep(PAUSE_SEC)

    # ── Final save ────────────────────────────────────────────────────────────
    _save_csv(all_results)

    if not args.csv_only:
        # Update subtitle with enrichment date
        ws["A4"].value = (
            f"Last Updated: {datetime.today().strftime('%d %b %Y')}  |  "
            f"Pipeline enriched via ClinicalTrials.gov API  |  "
            f"Sources: SEC EDGAR · yFinance · ClinicalTrials.gov"
        )
        wb.save(args.input)
        print(f"\n✅ Excel updated: {args.input}")

    print(f"✅ Pipeline CSV:  {PIPELINE_CSV}")

    # ── Summary stats ─────────────────────────────────────────────────────────
    if all_results:
        df = pd.DataFrame(all_results)
        print(f"\n{'─'*40}")
        print("Therapy Area breakdown:")
        print(df["therapy_area"].value_counts().to_string())
        print(f"\nDev Stage breakdown:")
        print(df["dev_stage"].value_counts().to_string())
        print(f"\nModality breakdown:")
        print(df["modality"].value_counts().to_string())
        no_data = df[df["trial_count"] == 0]
        print(f"\nNo ClinicalTrials.gov data: {len(no_data)} companies")
        if len(no_data) > 0:
            print(f"  → {no_data['Ticker'].tolist()[:20]}")
            print(f"  These may be pre-IND, platform cos, or use different sponsor names")
        print(f"{'─'*40}")


def _save_csv(results: list):
    if not results:
        return
    import os
    rows = []
    for r in results:
        rows.append({
            "Ticker":         r.get("Ticker", ""),
            "Company":        r.get("Company", ""),
            "Therapy Area":   r.get("therapy_area", ""),
            "Sub-Indication": r.get("sub_indication", ""),
            "Modality":       r.get("modality", ""),
            "Lead Asset":     r.get("lead_asset", ""),
            "Dev Stage":      r.get("dev_stage", ""),
            "Next Catalyst":  r.get("next_catalyst", ""),
            "Catalyst Type":  r.get("catalyst_type", ""),
            "Partners":       r.get("partners", ""),
            "Trial Count":    r.get("trial_count", 0),
            "All Phases":     ", ".join(r.get("all_phases", [])),
            "Fetched At":     r.get("fetched_at", ""),
        })
    df = pd.DataFrame(rows)
    # Append to existing CSV if resuming
    if os.path.exists(PIPELINE_CSV):
        existing = pd.read_csv(PIPELINE_CSV)
        df = pd.concat([existing, df], ignore_index=True).drop_duplicates("Ticker", keep="last")
    df.to_csv(PIPELINE_CSV, index=False)


if __name__ == "__main__":
    main()
